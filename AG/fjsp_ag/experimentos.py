
import os
import platform
import sys
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from fjsp_parser import leer_instancia
from ag_fjsp import resolver_ag



# CONFIGURACION DEL EXPERIMENTO
SEMILLA_BASE = 100
N_REPETICIONES = 5
PRESUPUESTO_SEG = 60.0

CARPETA_INSTANCIAS = "instancias"
CARPETA_SALIDA = "resultados"

# Cada entrada describe un benchmark: nombre de hoja/seccion, carpeta donde
# estan los archivos, extension, lista de instancias y sus Upper Bounds.
BENCHMARKS = {
    "Hurink_edata": {
        "ruta": os.path.join(CARPETA_INSTANCIAS, "hurink", "edata"),
        "extension": ".txt",
        "instancias": ["la01", "la06", "la11", "la16", "la21"],
        "ub": {"la01": 609, "la06": 800, "la11": 1071, "la16": 717, "la21": 835},
    },
    "Hurink_rdata": {
        "ruta": os.path.join(CARPETA_INSTANCIAS, "hurink", "rdata"),
        "extension": ".txt",
        "instancias": ["la01", "la06", "la11", "la16", "la21"],
        "ub": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 833},
    },
    "Hurink_vdata": {
        "ruta": os.path.join(CARPETA_INSTANCIAS, "hurink", "vdata"),
        "extension": ".txt",
        "instancias": ["la01", "la06", "la11", "la16", "la21"],
        "ub": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 800},
    },
    "Brandimarte": {
        "ruta": os.path.join(CARPETA_INSTANCIAS, "brandimarte"),
        "extension": ".fjs",
        "instancias": ["Mk01", "Mk02", "Mk03", "Mk04", "Mk05",
                       "Mk06", "Mk07", "Mk08", "Mk09", "Mk10"],
        "ub": {"Mk01": 42, "Mk02": 32, "Mk03": 211, "Mk04": 81, "Mk05": 186,
               "Mk06": 86, "Mk07": 157, "Mk08": 523, "Mk09": 369, "Mk10": 296},
    },
    "Dauzere": {
        "ruta": os.path.join(CARPETA_INSTANCIAS, "dauzere"),
        "extension": ".fjs",
        "instancias": ["01a", "04a", "07a", "10a", "13a", "16a"],
        "ub": {"01a": 2530, "04a": 2565, "07a": 2408,
               "10a": 2362, "13a": 2302, "16a": 2301},
    },
}

# Agrupacion para el reporte TXT (Hurink se presenta junto, con sus 3 subsets).
GRUPOS_REPORTE = [
    ("HURINK (edata / rdata / vdata)", ["Hurink_edata", "Hurink_rdata", "Hurink_vdata"]),
    ("BRANDIMARTE", ["Brandimarte"]),
    ("DAUZERE-PERES & PAULLI", ["Dauzere"]),
]

# Mapeo benchmark -> nombre de hoja "Rep_..." en el Excel.
HOJA_REP = {
    "Hurink_edata": "Rep_Hurink_edata",
    "Hurink_rdata": "Rep_Hurink_rdata",
    "Hurink_vdata": "Rep_Hurink_vdata",
    "Brandimarte": "Rep_Brandimarte",
    "Dauzere": "Rep_Dauzere",
}



# UTILIDADES ESTADISTICAS
def media(xs):
    return sum(xs) / len(xs)


def desv_std_muestral(xs):
    n = len(xs)
    if n < 2:
        return 0.0
    m = media(xs)
    return (sum((x - m) ** 2 for x in xs) / (n - 1)) ** 0.5



# DETECCION DE HARDWARE / ENTORNO (solo biblioteca estandar)
def detectar_entorno():
    info = {}
    info["sistema_operativo"] = "{} {} ({})".format(
        platform.system(), platform.release(), platform.version())
    info["arquitectura"] = platform.machine()
    info["procesador"] = platform.processor() or "no identificado por el SO"
    try:
        info["nucleos_logicos"] = str(os.cpu_count())
    except Exception:
        info["nucleos_logicos"] = "desconocido"
    info["python_version"] = platform.python_version()
    info["implementacion_python"] = platform.python_implementation()
    return info

# EJECUCION DE UN BLOQUE DE REPETICIONES PARA UNA INSTANCIA
def correr_instancia(benchmark, nombre, presupuesto):
    cfg = BENCHMARKS[benchmark]
    ruta = os.path.join(cfg["ruta"], nombre + cfg["extension"])
    etiqueta = benchmark + "/" + nombre
    inst = leer_instancia(ruta, etiqueta)
    ub = cfg["ub"][nombre]

    detalle = []
    makespans = []
    generaciones = []
    tiempos = []
    evals = []

    for r in range(N_REPETICIONES):
        semilla = SEMILLA_BASE + r
        res = resolver_ag(inst, semilla, presupuesto)
        mk = res["makespan"]
        gap = (mk - ub) / ub * 100.0

        makespans.append(mk)
        generaciones.append(res["generaciones"])
        tiempos.append(res["tiempo"])
        evals.append(res["evaluaciones"])

        fila = {
            "benchmark": benchmark,
            "instancia": nombre,
            "repeticion": r + 1,
            "semilla": semilla,
            "UB": ub,
            "makespan": mk,
            "gap_%": round(gap, 4),
            "generaciones": res["generaciones"],
            "evaluaciones": res["evaluaciones"],
            "tiempo_s": round(res["tiempo"], 3),
        }
        detalle.append(fila)
        print("  [{}] sem={} -> makespan={}  gap={:.2f}%  gen={}  t={:.1f}s"
              .format(etiqueta, semilla, mk, gap, res["generaciones"], res["tiempo"]))

    mejor = min(makespans)
    peor = max(makespans)
    prom = media(makespans)
    std = desv_std_muestral(makespans)

    resumen = {
        "benchmark": benchmark,
        "instancia": nombre,
        "UB": ub,
        "mejor": mejor,
        "promedio": round(prom, 3),
        "desv_std": round(std, 3),
        "peor": peor,
        "gap_mejor_%": round((mejor - ub) / ub * 100.0, 4),
        "gap_promedio_%": round((prom - ub) / ub * 100.0, 4),
        "iter_promedio": round(media(generaciones), 1),
        "evals_promedio": round(media(evals), 1),
        "tiempo_promedio_s": round(media(tiempos), 3),
    }
    return detalle, resumen


# EXPORTACION: configuracion.txt
def exportar_configuracion(ruta, presupuesto, ts_legible, duracion_total_s):
    entorno = detectar_entorno()
    semillas = [SEMILLA_BASE + r for r in range(N_REPETICIONES)]
    total_instancias = sum(len(c["instancias"]) for c in BENCHMARKS.values())
    total_ejecuciones = total_instancias * N_REPETICIONES

    lineas = []
    lineas.append("=" * 70)
    lineas.append("CONFIGURACION DEL EXPERIMENTO")
    lineas.append("=" * 70)
    lineas.append("Fecha/hora de la corrida : " + ts_legible)
    lineas.append("Duracion total           : {:.1f} s ({:.2f} min)".format(
        duracion_total_s, duracion_total_s / 60.0))
    lineas.append("")
    lineas.append("Metaheuristica            : Algoritmo Genetico (AG)")
    lineas.append("Repeticiones por instancia: {}".format(N_REPETICIONES))
    lineas.append("Semillas (base {})        : {}".format(SEMILLA_BASE, semillas))
    lineas.append("Presupuesto por ejecucion : {:.0f} segundos (uniforme)".format(presupuesto))
    lineas.append("Total de instancias       : {}".format(total_instancias))
    lineas.append("Total de ejecuciones      : {}".format(total_ejecuciones))
    lineas.append("")
    lineas.append("Benchmarks incluidos:")
    for nombre_b, cfg in BENCHMARKS.items():
        lineas.append("  - {:<14}: {} instancias -> {}".format(
            nombre_b, len(cfg["instancias"]), ", ".join(cfg["instancias"])))
    lineas.append("")
    lineas.append("-" * 70)
    lineas.append("ENTORNO DE EJECUCION (detectado automaticamente)")
    lineas.append("-" * 70)
    lineas.append("Sistema operativo  : " + entorno["sistema_operativo"])
    lineas.append("Arquitectura       : " + entorno["arquitectura"])
    lineas.append("Procesador         : " + entorno["procesador"])
    lineas.append("Nucleos logicos    : " + entorno["nucleos_logicos"])
    lineas.append("Python             : {} ({})".format(
        entorno["python_version"], entorno["implementacion_python"]))
    lineas.append("=" * 70)

    with open(ruta, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas) + "\n")


# EXPORTACION: resultados_reporte.txt
def _tabla_resumen_txt(resumenes_benchmark):
    cab = ("{:<8} {:>6} {:>6} {:>9} {:>8} {:>6} {:>10} {:>11} {:>9} {:>9}"
           .format("inst", "UB", "mejor", "promedio", "desvstd",
                   "peor", "gap_mej_%", "gap_prom_%", "iter_pr", "t_prom_s"))
    lineas = [cab, "-" * len(cab)]
    for r in resumenes_benchmark:
        lineas.append("{:<8} {:>6} {:>6} {:>9} {:>8} {:>6} {:>10} {:>11} {:>9} {:>9}"
                      .format(r["instancia"], r["UB"], r["mejor"], r["promedio"],
                              r["desv_std"], r["peor"], r["gap_mejor_%"],
                              r["gap_promedio_%"], r["iter_promedio"],
                              r["tiempo_promedio_s"]))
    return lineas


def exportar_reporte_txt(ruta, resumenes_por_benchmark, ts_legible, presupuesto):
    lineas = []
    lineas.append("=" * 78)
    lineas.append("  RESULTADOS - ALGORITMO GENETICO PARA FJSP")
    lineas.append("  Objetivo: minimizar makespan (Cmax)")
    lineas.append("=" * 78)
    lineas.append("Fecha/hora        : " + ts_legible)
    lineas.append("Repeticiones      : {} (semillas {}..{})".format(
        N_REPETICIONES, SEMILLA_BASE, SEMILLA_BASE + N_REPETICIONES - 1))
    lineas.append("Presupuesto/ejec. : {:.0f} segundos (uniforme)".format(presupuesto))
    lineas.append("")

    todos_gaps_mejor = []
    for titulo_grupo, benchmarks_grupo in GRUPOS_REPORTE:
        lineas.append("#" * 78)
        lineas.append("  " + titulo_grupo)
        lineas.append("#" * 78)
        for nombre_b in benchmarks_grupo:
            resumenes_b = resumenes_por_benchmark[nombre_b]
            lineas.append("")
            lineas.append("--- " + nombre_b + " ---")
            lineas.extend(_tabla_resumen_txt(resumenes_b))
            gaps = [r["gap_mejor_%"] for r in resumenes_b]
            todos_gaps_mejor.extend(gaps)
            lineas.append("Gap promedio (del mejor) en {}: {:.2f}%".format(
                nombre_b, media(gaps)))
        lineas.append("")

    lineas.append("=" * 78)
    lineas.append("GAP PROMEDIO GLOBAL (del mejor, sobre todas las instancias): {:.2f}%"
                  .format(media(todos_gaps_mejor)))
    lineas.append("=" * 78)

    with open(ruta, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas) + "\n")



# EXPORTACION: resultados.xlsx
COLOR_ENCABEZADO = "1F4E78"
FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)
RELLENO_ENCABEZADO = PatternFill(start_color=COLOR_ENCABEZADO,
                                  end_color=COLOR_ENCABEZADO, fill_type="solid")


def _escribir_tabla(ws, columnas, filas, fila_inicio=1):
    for c, nombre_col in enumerate(columnas, start=1):
        celda = ws.cell(row=fila_inicio, column=c, value=nombre_col)
        celda.font = FUENTE_ENCABEZADO
        celda.fill = RELLENO_ENCABEZADO
        celda.alignment = Alignment(horizontal="center")
    for i, fila in enumerate(filas, start=1):
        for c, nombre_col in enumerate(columnas, start=1):
            ws.cell(row=fila_inicio + i, column=c, value=fila.get(nombre_col))
    for c, nombre_col in enumerate(columnas, start=1):
        ancho = max(12, len(str(nombre_col)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = ancho
    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)


COL_RESUMEN = ["benchmark", "instancia", "UB", "mejor", "promedio", "desv_std",
               "peor", "gap_mejor_%", "gap_promedio_%", "iter_promedio",
               "evals_promedio", "tiempo_promedio_s"]
COL_DETALLE = ["benchmark", "instancia", "repeticion", "semilla", "UB",
               "makespan", "gap_%", "generaciones", "evaluaciones", "tiempo_s"]


def exportar_excel(ruta, resumenes_por_benchmark, detalles_por_benchmark):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Resumen_General"
    todas_resumenes = []
    for nombre_b in BENCHMARKS:
        todas_resumenes.extend(resumenes_por_benchmark[nombre_b])
    columnas_sin_bench = [c for c in COL_RESUMEN if c != "benchmark"]
    _escribir_tabla(ws0, COL_RESUMEN, todas_resumenes)

    hojas_benchmark = {
        "Hurink_edata": "Hurink_edata",
        "Hurink_rdata": "Hurink_rdata",
        "Hurink_vdata": "Hurink_vdata",
        "Brandimarte": "Brandimarte",
        "Dauzere": "Dauzere",
    }
    for nombre_b, nombre_hoja in hojas_benchmark.items():
        ws = wb.create_sheet(nombre_hoja)
        _escribir_tabla(ws, columnas_sin_bench,
                        [{k: v for k, v in r.items() if k != "benchmark"}
                         for r in resumenes_por_benchmark[nombre_b]])

    columnas_detalle_sin_bench = [c for c in COL_DETALLE if c != "benchmark"]
    for nombre_b, nombre_hoja in HOJA_REP.items():
        ws = wb.create_sheet(nombre_hoja)
        _escribir_tabla(ws, columnas_detalle_sin_bench,
                        [{k: v for k, v in d.items() if k != "benchmark"}
                         for d in detalles_por_benchmark[nombre_b]])

    wb.save(ruta)



# PROGRAMA PRINCIPAL
def main():
    presupuesto = PRESUPUESTO_SEG
    if "--rapido" in sys.argv:
        presupuesto = 3.0
        print(">> MODO RAPIDO: presupuesto reducido a {:.0f}s por ejecucion.\n"
              .format(presupuesto))

    os.makedirs(CARPETA_SALIDA, exist_ok=True)
    t_inicio_total = time.perf_counter()
    ts_legible = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    detalles_por_benchmark = {nombre_b: [] for nombre_b in BENCHMARKS}
    resumenes_por_benchmark = {nombre_b: [] for nombre_b in BENCHMARKS}

    n_total = sum(len(c["instancias"]) for c in BENCHMARKS.values())
    contador = 0
    for nombre_b, cfg in BENCHMARKS.items():
        for inst_nom in cfg["instancias"]:
            contador += 1
            print("[{}/{}] Resolviendo {}/{} ..."
                  .format(contador, n_total, nombre_b, inst_nom))
            det, res = correr_instancia(nombre_b, inst_nom, presupuesto)
            detalles_por_benchmark[nombre_b].extend(det)
            resumenes_por_benchmark[nombre_b].append(res)
            print("    -> mejor={}  prom={}  std={}  peor={}  gap_mejor={:.2f}%\n"
                  .format(res["mejor"], res["promedio"], res["desv_std"],
                          res["peor"], res["gap_mejor_%"]))

    duracion_total = time.perf_counter() - t_inicio_total

    ruta_txt = os.path.join(CARPETA_SALIDA, "resultados_reporte.txt")
    ruta_xlsx = os.path.join(CARPETA_SALIDA, "resultados.xlsx")
    ruta_config = os.path.join(CARPETA_SALIDA, "configuracion.txt")

    exportar_reporte_txt(ruta_txt, resumenes_por_benchmark, ts_legible, presupuesto)
    exportar_excel(ruta_xlsx, resumenes_por_benchmark, detalles_por_benchmark)
    exportar_configuracion(ruta_config, presupuesto, ts_legible, duracion_total)

    print("Listo. Archivos generados en '{}':".format(CARPETA_SALIDA))
    print("  -", ruta_txt)
    print("  -", ruta_xlsx)
    print("  -", ruta_config)


if __name__ == "__main__":
    main()
