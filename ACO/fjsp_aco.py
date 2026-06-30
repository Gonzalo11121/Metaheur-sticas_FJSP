import argparse
import math
import os
import platform
import random
import statistics
import time
from datetime import datetime



#bounds y configuracion de benchmarks
HURINK_INSTANCES = ["la01", "la06", "la11", "la16", "la21"]

UB_HURINK = {
    "edata": {"la01": 609, "la06": 800, "la11": 1071, "la16": 717, "la21": 835},
    "rdata": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 833},
    "vdata": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 800},
}

UB_BRANDIMARTE = {
    "Mk01": 42,  "Mk02": 32,  "Mk03": 211, "Mk04": 81,  "Mk05": 186,
    "Mk06": 86,  "Mk07": 157, "Mk08": 523, "Mk09": 369, "Mk10": 296,
}

UB_DAUZERE = {
    "01a": 2530, "04a": 2565, "07a": 2408,
    "10a": 2362, "13a": 2302, "16a": 2301,
}

BENCHMARKS = {
    "Hurink": {
        "groups": ["edata", "rdata", "vdata"],
        "subdir": "Hurink",
        "ext": ".txt",
        "group_subdir": True,
        "instances": {g: HURINK_INSTANCES for g in ["edata", "rdata", "vdata"]},
        "ub": UB_HURINK,
    },
    "Brandimarte": {
        "groups": ["Brandimarte"],
        "subdir": "Brandimarte",
        "ext": ".fjs",
        "group_subdir": False,
        "instances": {"Brandimarte": list(UB_BRANDIMARTE.keys())},
        "ub": {"Brandimarte": UB_BRANDIMARTE},
    },
    "Dauzere": {
        "groups": ["Dauzere"],
        "subdir": "Dauzere",
        "ext": ".fjs",
        "group_subdir": False,
        "instances": {"Dauzere": list(UB_DAUZERE.keys())},
        "ub": {"Dauzere": UB_DAUZERE},
    },
}


def ruta_instancia(base, benchmark, group, inst_name):
    cfg = BENCHMARKS[benchmark]
    if cfg["group_subdir"]:
        return os.path.join(base, cfg["subdir"], group, inst_name + cfg["ext"])
    return os.path.join(base, cfg["subdir"], inst_name + cfg["ext"])



# Lectura de instancias
def parse_fjs(path):
    with open(path, "r") as f:
        contenido = f.read().split("\n")
    lineas = [ln.strip() for ln in contenido if ln.strip() != ""]
    cabecera = lineas[0].split()
    n_jobs = int(cabecera[0])
    n_machines = int(cabecera[1])

    alts, op_id, job_of_op = [], [], []
    contador = 0
    for j in range(n_jobs):
        tokens = lineas[1 + j].split()
        idx = 0
        n_ops = int(tokens[idx]); idx += 1
        fila_ids = []
        for o in range(n_ops):
            n_maq = int(tokens[idx]); idx += 1
            opciones = []
            for _ in range(n_maq):
                m = int(tokens[idx]) - 1
                p = int(tokens[idx + 1])
                idx += 2
                opciones.append((m, p))
            alts.append(opciones)
            fila_ids.append(contador)
            job_of_op.append((j, o))
            contador += 1
        op_id.append(fila_ids)

    return {
        "name": os.path.basename(path),
        "n_jobs": n_jobs, "n_machines": n_machines,
        "n_ops": contador, "alts": alts,
        "op_id": op_id, "job_of_op": job_of_op,
    }



# Decodificador
def decode(inst, secuencia):
    libre_maquina = [0] * inst["n_machines"]
    listo_job     = [0] * inst["n_jobs"]
    makespan = 0
    for (j, o, m) in secuencia:
        g = inst["op_id"][j][o]
        p = next(pp for (mm, pp) in inst["alts"][g] if mm == m)
        inicio = max(libre_maquina[m], listo_job[j])
        fin = inicio + p
        libre_maquina[m] = fin
        listo_job[j]     = fin
        if fin > makespan:
            makespan = fin
    return makespan


# ACO - MAX-MIN 
class AntColonyMMAS:

    def __init__(self, inst, n_ants=30, alpha=1.0, beta=2.0,
                 rho=0.1, q0=0.0, local_search=True):
        self.inst = inst
        self.n_ants = n_ants
        self.alpha  = alpha
        self.beta   = beta
        self.rho    = rho
        self.q0     = q0
        self.local_search = local_search

        self.n_ops      = inst["n_ops"]
        self.n_machines = inst["n_machines"]
        self.n_jobs     = inst["n_jobs"]
        self.n_ops_job  = [len(r) for r in inst["op_id"]]

        self.tau = [[0.0] * self.n_machines for _ in range(self.n_ops)]
        self.celdas_validas = []
        for g in range(self.n_ops):
            for (m, _) in inst["alts"][g]:
                self.tau[g][m] = 1.0
                self.celdas_validas.append((g, m))

        self.mejor_makespan  = math.inf
        self.mejor_secuencia = None
        self.tau_max = 1.0
        self.tau_min = 0.0
        self.iteraciones = 0
        self.historial = []          # (iteracion, tiempo_s, mejor_makespan)

    def construir(self):
        inst       = self.inst
        prox_op    = [0] * self.n_jobs
        libre_maq  = [0] * self.n_machines
        listo_job  = [0] * self.n_jobs
        secuencia  = []

        for _ in range(self.n_ops):
            candidatos, pesos = [], []
            for j in range(self.n_jobs):
                o = prox_op[j]
                if o >= self.n_ops_job[j]:
                    continue
                g = inst["op_id"][j][o]
                for (m, p) in inst["alts"][g]:
                    fin = max(libre_maq[m], listo_job[j]) + p
                    eta = 1.0 / fin if fin > 0 else 1.0
                    candidatos.append((j, o, m, fin))
                    pesos.append((self.tau[g][m] ** self.alpha) *
                                 (eta ** self.beta))

            total = sum(pesos)
            if total <= 0:
                pesos = [1.0] * len(pesos)
                total = float(len(pesos))

            if self.q0 > 0 and random.random() < self.q0:
                k = max(range(len(candidatos)), key=lambda i: pesos[i])
            else:
                r, acum = random.random() * total, 0.0
                k = len(candidatos) - 1
                for i, w in enumerate(pesos):
                    acum += w
                    if acum >= r:
                        k = i; break

            j, o, m, fin = candidatos[k]
            libre_maq[m] = fin
            listo_job[j] = fin
            prox_op[j]  += 1
            secuencia.append((j, o, m))

        return max(listo_job), secuencia

    def busqueda_local(self, secuencia, deadline):
        mejor_seq = list(secuencia)
        mejor_mk  = decode(self.inst, mejor_seq)
        mejoro    = True
        while mejoro and time.time() < deadline:
            mejoro = False
            for i in range(len(mejor_seq)):
                if time.time() >= deadline:
                    break
                j, o, m_actual = mejor_seq[i]
                g = self.inst["op_id"][j][o]
                for (m, _) in self.inst["alts"][g]:
                    if m == m_actual:
                        continue
                    prueba = list(mejor_seq)
                    prueba[i] = (j, o, m)
                    mk = decode(self.inst, prueba)
                    if mk < mejor_mk:
                        mejor_mk, mejor_seq, mejoro = mk, prueba, True
                        break
        return mejor_mk, mejor_seq

    def actualizar_feromona(self):
        for (g, m) in self.celdas_validas:
            self.tau[g][m] *= (1.0 - self.rho)
        dep = 1.0 / self.mejor_makespan
        for (j, o, m) in self.mejor_secuencia:
            self.tau[self.inst["op_id"][j][o]][m] += dep
        self.tau_max = 1.0 / (self.rho * self.mejor_makespan)
        self.tau_min = self.tau_max / (2.0 * self.n_ops)
        for (g, m) in self.celdas_validas:
            self.tau[g][m] = max(self.tau_min,
                                 min(self.tau_max, self.tau[g][m]))

    def resolver(self, time_budget, seed=None, guardar_historial=False):
        if seed is not None:
            random.seed(seed)
        t0       = time.time()
        deadline = t0 + time_budget
        self.iteraciones = 0
        self.historial   = []
        prev_best        = math.inf

        while time.time() < deadline:
            iter_mk, iter_seq = math.inf, None
            for _ in range(self.n_ants):
                mk, seq = self.construir()
                if mk < iter_mk:
                    iter_mk, iter_seq = mk, seq
                if time.time() >= deadline:
                    break

            nuevo_mejor = False
            if iter_seq is not None and iter_mk < self.mejor_makespan:
                self.mejor_makespan  = iter_mk
                self.mejor_secuencia = iter_seq
                nuevo_mejor = True

            if self.local_search and nuevo_mejor and time.time() < deadline:
                mk2, seq2 = self.busqueda_local(self.mejor_secuencia, deadline)
                if mk2 < self.mejor_makespan:
                    self.mejor_makespan  = mk2
                    self.mejor_secuencia = seq2

            if self.mejor_secuencia is not None:
                self.actualizar_feromona()

            self.iteraciones += 1

            if guardar_historial and self.mejor_makespan < prev_best:
                self.historial.append(
                    (self.iteraciones, round(time.time() - t0, 3),
                     self.mejor_makespan))
                prev_best = self.mejor_makespan

        tiempo_real = time.time() - t0
        if guardar_historial:
            self.historial.append(
                (self.iteraciones, round(tiempo_real, 3), self.mejor_makespan))
        return self.mejor_makespan, self.iteraciones, tiempo_real



# Hardware
def detectar_hardware():
    info = {
        "sistema":        f"{platform.system()} {platform.release()}",
        "arquitectura":   platform.machine(),
        "procesador":     platform.processor() or "no disponible",
        "python":         platform.python_version(),
        "nucleos_logicos": os.cpu_count() or "no disponible",
        "nucleos_fisicos": "no disponible",
        "ram_total_gb":   "no disponible",
    }
    try:
        import psutil
        info["nucleos_fisicos"] = psutil.cpu_count(logical=False)
        info["ram_total_gb"]    = round(
            psutil.virtual_memory().total / (1024 ** 3), 2)
        frec = psutil.cpu_freq()
        if frec:
            info["frecuencia_mhz"] = round(frec.max or frec.current, 0)
    except Exception:
        pass
    return info



# Experimento
def correr_experimento(args):
    os.makedirs(args.out, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    semillas  = [args.seed_base + r for r in range(args.reps)]

    resumen, detalle, convergencia = [], [], []

    print("=" * 72)
    print(" EXPERIMENTO ACO (MMAS) - FJSP")
    print(f" benchmarks={args.benchmarks}  reps={args.reps}  "
          f"semillas={semillas}  presupuesto={args.budget}s/corrida")
    print("=" * 72)

    for benchmark in args.benchmarks:
        cfg = BENCHMARKS[benchmark]
        for group in cfg["groups"]:
            for inst_name in cfg["instances"][group]:
                ruta = ruta_instancia(
                    args.instances_dir, benchmark, group, inst_name)
                if not os.path.isfile(ruta):
                    print(f"  [AVISO] no se encontro {ruta}, se omite.")
                    continue

                inst = parse_fjs(ruta)
                ub   = cfg["ub"][group].get(inst_name)
                etiq = (f"{benchmark}/{group}/{inst_name}"
                        if cfg["group_subdir"] else f"{benchmark}/{inst_name}")
                print(f"\n>>> {etiq}  ({inst['n_jobs']}x{inst['n_machines']}, "
                      f"{inst['n_ops']} ops)  UB={ub}")

                makespans, iters, tiempos = [], [], []

                for rep in range(args.reps):
                    seed           = semillas[rep]
                    es_primera_rep = (rep == 0)
                    aco = AntColonyMMAS(
                        inst, n_ants=args.ants, alpha=args.alpha,
                        beta=args.beta, rho=args.rho, q0=args.q0,
                        local_search=not args.no_localsearch)
                    mk, it, t = aco.resolver(
                        args.budget, seed=seed,
                        guardar_historial=es_primera_rep)

                    makespans.append(mk)
                    iters.append(it)
                    tiempos.append(t)
                    gap_run = (mk - ub) / ub * 100 if ub else None

                    if ub:
                        print(f"    rep {rep+1}/{args.reps}  seed={seed}  "
                              f"Cmax={mk}  gap={gap_run:6.2f}%  "
                              f"iters={it}  t={t:.1f}s")
                    else:
                        print(f"    rep {rep+1}/{args.reps}  seed={seed}  "
                              f"Cmax={mk}  iters={it}  t={t:.1f}s")

                    detalle.append({
                        "benchmark": benchmark, "grupo": group,
                        "instancia": inst_name, "rep": rep + 1, "seed": seed,
                        "makespan": mk,
                        "gap_pct": round(gap_run, 3) if ub else "",
                        "iteraciones": it, "tiempo_s": round(t, 2),
                    })

                    # convergencia: solo rep 1 (seed base)
                    if es_primera_rep:
                        for (it_conv, t_conv, mk_conv) in aco.historial:
                            convergencia.append({
                                "benchmark": benchmark, "grupo": group,
                                "instancia": inst_name, "seed": seed,
                                "iteracion": it_conv,
                                "tiempo_s":  t_conv,
                                "mejor_makespan": mk_conv,
                            })

                mejor    = min(makespans)
                peor     = max(makespans)
                promedio = statistics.mean(makespans)
                desv     = (statistics.stdev(makespans)
                            if len(makespans) > 1 else 0.0)
                gap      = (mejor - ub) / ub * 100 if ub else None

                resumen.append({
                    "benchmark": benchmark, "grupo": group,
                    "instancia": inst_name,
                    "n_jobs": inst["n_jobs"], "n_machines": inst["n_machines"],
                    "n_ops": inst["n_ops"], "UB": ub,
                    "mejor": mejor, "promedio": round(promedio, 2),
                    "desv": round(desv, 2), "peor": peor,
                    "gap_pct": round(gap, 2) if ub else "",
                    "iter_prom": round(statistics.mean(iters), 1),
                    "tiempo_prom_s": round(statistics.mean(tiempos), 2),
                    "reps": args.reps,
                })
                if ub:
                    print(f"    => mejor={mejor}  prom={promedio:.1f}  "
                          f"desv={desv:.2f}  peor={peor}  gap={gap:.2f}%")
                else:
                    print(f"    => mejor={mejor}  prom={promedio:.1f}  peor={peor}")

    exportar_txt(args, resumen, timestamp)
    exportar_config(args, semillas, timestamp)
    exportar_xlsx(args, resumen, detalle, convergencia)

    print("\n" + "=" * 72)
    print(" ARCHIVOS GENERADOS en:", args.out)
    print("   resultados.xlsx  (todas las hojas en un solo archivo)")
    print("   resultados_reporte.txt")
    print("   configuracion.txt")
    print("=" * 72)



# Exportacion TXT
def exportar_txt(args, resumen, timestamp):
    ruta = os.path.join(args.out, "resultados_reporte.txt")
    cols = (f"{'instancia':10} {'UB':>6} {'mejor':>7} {'promedio':>9} "
            f"{'desv':>7} {'peor':>7} {'gap%':>7} {'iter':>8} {'t(s)':>7}")
    with open(ruta, "w", encoding="utf-8") as f:
        f.write("RESULTADOS - ACO (MAX-MIN Ant System) - FJSP\n")
        f.write(f"Fecha: {timestamp}\n")
        f.write("gap% = (mejor - UB) / UB * 100\n")
        f.write("=" * 78 + "\n")
        for benchmark in args.benchmarks:
            cfg = BENCHMARKS[benchmark]
            for group in cfg["groups"]:
                filas = [r for r in resumen
                         if r["benchmark"] == benchmark
                         and r["grupo"] == group]
                if not filas:
                    continue
                titulo = (f"{benchmark} - {group}"
                          if cfg["group_subdir"] else benchmark)
                f.write(f"\n{titulo}\n" + "-" * 78 + "\n" + cols + "\n")
                gaps = []
                for r in filas:
                    f.write(f"{r['instancia']:10} {str(r['UB']):>6} "
                            f"{r['mejor']:>7} {r['promedio']:>9} "
                            f"{r['desv']:>7} {r['peor']:>7} "
                            f"{str(r['gap_pct']):>7} "
                            f"{r['iter_prom']:>8} {r['tiempo_prom_s']:>7}\n")
                    if r["gap_pct"] != "":
                        gaps.append(r["gap_pct"])
                if gaps:
                    f.write(f"{'gap% promedio:':>54} "
                            f"{statistics.mean(gaps):6.2f}\n")
        f.write("\n" + "=" * 78 + "\n")



# Exportacion configuracion

def exportar_config(args, semillas, timestamp):
    hw   = detectar_hardware()
    ruta = os.path.join(args.out, "configuracion.txt")
    with open(ruta, "w", encoding="utf-8") as f:
        f.write("CONFIGURACION DEL EXPERIMENTO\n" + "=" * 50 + "\n")
        f.write(f"Fecha              : {timestamp}\n\n")
        f.write("Metaheuristica\n" + "-" * 50 + "\n")
        f.write("  Algoritmo        : Colonia de Hormigas (ACO)\n")
        f.write("  Variante         : MAX-MIN Ant System (MMAS)\n")
        f.write("  Objetivo         : minimizar makespan (Cmax)\n\n")
        f.write("Parametros\n" + "-" * 50 + "\n")
        f.write(f"  Hormigas         : {args.ants}\n")
        f.write(f"  alpha            : {args.alpha}\n")
        f.write(f"  beta             : {args.beta}\n")
        f.write(f"  rho              : {args.rho}\n")
        f.write(f"  q0               : {args.q0}\n")
        f.write(f"  Busqueda local   : {'no' if args.no_localsearch else 'si'}\n")
        f.write(f"  Presupuesto/rep  : {args.budget} s\n")
        f.write(f"  Repeticiones     : {args.reps}\n")
        f.write(f"  Semillas         : {semillas}\n\n")
        f.write("Benchmarks\n" + "-" * 50 + "\n")
        for b in args.benchmarks:
            cfg   = BENCHMARKS[b]
            total = sum(len(cfg["instances"][g]) for g in cfg["groups"])
            f.write(f"  {b:12}: {total} instancias "
                    f"({', '.join(cfg['groups'])})\n")
        f.write("\nHardware\n" + "-" * 50 + "\n")
        f.write(f"  Sistema          : {hw['sistema']}\n")
        f.write(f"  Arquitectura     : {hw['arquitectura']}\n")
        f.write(f"  Procesador       : {hw['procesador']}\n")
        f.write(f"  Nucleos fisicos  : {hw['nucleos_fisicos']}\n")
        f.write(f"  Nucleos logicos  : {hw['nucleos_logicos']}\n")
        if "frecuencia_mhz" in hw:
            f.write(f"  Frecuencia (MHz) : {hw['frecuencia_mhz']}\n")
        f.write(f"  RAM (GB)         : {hw['ram_total_gb']}\n")
        f.write(f"  Python           : {hw['python']}\n")



# Exportacion Excel:

def exportar_xlsx(args, resumen, detalle, convergencia):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [AVISO] openpyxl no instalado. "
              "Ejecuta: pip install openpyxl")
        return

    ruta = os.path.join(args.out, "resultados.xlsx")

    enc_res = ["Benchmark", "Grupo", "Instancia", "Jobs", "Maquinas",
               "Operaciones", "UB", "Mejor", "Promedio", "DesvEst",
               "Peor", "Gap%", "IterProm", "TiempoPromS", "Reps"]
    enc_rep = ["Benchmark", "Grupo", "Instancia", "Rep", "Seed",
               "Makespan", "Gap%", "Iteraciones", "TiempoS"]
    enc_conv = ["Benchmark", "Grupo", "Instancia", "Seed",
                "Iteracion", "TiempoS", "MejorMakespan"]

    h_tit  = Font(name="Arial", bold=True, color="FFFFFF")
    fill_b = PatternFill("solid", start_color="305496")   # azul oscuro
    fill_g = PatternFill("solid", start_color="375623")   # verde oscuro
    fill_r = PatternFill("solid", start_color="833C00")   # naranja oscuro
    fill_p = PatternFill("solid", start_color="4B0082")   # morado oscuro
    centro = Alignment(horizontal="center")


    color_por_hoja = {
        "Resumen_General":  fill_b,
        "Hurink_edata":     fill_b, "Hurink_rdata": fill_b, "Hurink_vdata": fill_b,
        "Brandimarte":      fill_g, "Dauzere": fill_r,
        "Rep_Hurink_edata": fill_b, "Rep_Hurink_rdata": fill_b,
        "Rep_Hurink_vdata": fill_b,
        "Rep_Brandimarte":  fill_g, "Rep_Dauzere": fill_r,
        "Convergencia":     fill_p,
    }

    def f_res(r):
        return [r["benchmark"], r["grupo"], r["instancia"],
                r["n_jobs"], r["n_machines"], r["n_ops"], r["UB"],
                r["mejor"], r["promedio"], r["desv"], r["peor"],
                r["gap_pct"], r["iter_prom"], r["tiempo_prom_s"], r["reps"]]

    def f_rep(r):
        return [r["benchmark"], r["grupo"], r["instancia"],
                r["rep"], r["seed"], r["makespan"],
                r["gap_pct"], r["iteraciones"], r["tiempo_s"]]

    def f_conv(r):
        return [r["benchmark"], r["grupo"], r["instancia"],
                r["seed"], r["iteracion"], r["tiempo_s"], r["mejor_makespan"]]

    def escribir_hoja(ws, encabezados, filas, fill):
        ws.append(encabezados)
        for c in range(1, len(encabezados) + 1):
            cel = ws.cell(row=1, column=c)
            cel.font      = h_tit
            cel.fill      = fill
            cel.alignment = centro
        for fila in filas:
            ws.append(fila)
        for c in range(1, len(encabezados) + 1):
            largos = ([len(str(encabezados[c - 1]))] +
                      [len(str(f[c - 1])) for f in filas])
            ws.column_dimensions[get_column_letter(c)].width = min(
                max(largos) + 2, 26)
        ws.freeze_panes = "A2"

    wb = Workbook()

    # Resumen_General
    ws = wb.active
    ws.title = "Resumen_General"
    escribir_hoja(ws, enc_res, [f_res(r) for r in resumen],
                  color_por_hoja["Resumen_General"])

    # hojas de resumen por grupo
    grupos_hojas = [
        ("Hurink",      "edata",       "Hurink_edata"),
        ("Hurink",      "rdata",       "Hurink_rdata"),
        ("Hurink",      "vdata",       "Hurink_vdata"),
        ("Brandimarte", "Brandimarte", "Brandimarte"),
        ("Dauzere",     "Dauzere",     "Dauzere"),
    ]
    for bm, gr, nombre in grupos_hojas:
        filas = [f_res(r) for r in resumen
                 if r["benchmark"] == bm and r["grupo"] == gr]
        escribir_hoja(wb.create_sheet(nombre), enc_res, filas,
                      color_por_hoja[nombre])

    # hojas de detalle (repeticiones)
    rep_hojas = [
        ("Hurink",      "edata",       "Rep_Hurink_edata"),
        ("Hurink",      "rdata",       "Rep_Hurink_rdata"),
        ("Hurink",      "vdata",       "Rep_Hurink_vdata"),
        ("Brandimarte", "Brandimarte", "Rep_Brandimarte"),
        ("Dauzere",     "Dauzere",     "Rep_Dauzere"),
    ]
    for bm, gr, nombre in rep_hojas:
        filas = [f_rep(r) for r in detalle
                 if r["benchmark"] == bm and r["grupo"] == gr]
        escribir_hoja(wb.create_sheet(nombre), enc_rep, filas,
                      color_por_hoja[nombre])

    # hoja Convergencia (todas las instancias juntas)
    ws_conv = wb.create_sheet("Convergencia")
    escribir_hoja(ws_conv, enc_conv, [f_conv(r) for r in convergencia],
                  color_por_hoja["Convergencia"])

    wb.save(ruta)
    print(f"  Excel guardado: {ruta}  "
          f"({len(wb.sheetnames)} hojas, "
          f"{len(convergencia)} filas de convergencia)")



# Interfaz de linea de comando

def main():
    ap = argparse.ArgumentParser(
        description="ACO (MMAS) para FJSP - Hurink, Brandimarte y Dauzere.")
    ap.add_argument("--instances-dir", default="instancias")
    ap.add_argument("--benchmarks", nargs="+",
                    default=["Hurink", "Brandimarte", "Dauzere"],
                    choices=["Hurink", "Brandimarte", "Dauzere"])
    ap.add_argument("--reps",       type=int,   default=5)
    ap.add_argument("--seed-base",  type=int,   default=100)
    ap.add_argument("--budget",     type=float, default=60.0)
    ap.add_argument("--ants",       type=int,   default=30)
    ap.add_argument("--alpha",      type=float, default=1.0)
    ap.add_argument("--beta",       type=float, default=2.0)
    ap.add_argument("--rho",        type=float, default=0.1)
    ap.add_argument("--q0",         type=float, default=0.0)
    ap.add_argument("--no-localsearch", action="store_true")
    ap.add_argument("--out",        default="resultados")
    args = ap.parse_args()
    correr_experimento(args)


if __name__ == "__main__":
    main()
