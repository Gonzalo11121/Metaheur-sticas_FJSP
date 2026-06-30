#Genera 11 gráficos necesarios para el experimento

import argparse
import glob
import os
from collections import defaultdict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook

SUBSETS_HURINK  = ["edata", "rdata", "vdata"]
INSTANCIAS_HURINK = ["la01", "la06", "la11", "la16", "la21"]
INSTANCIAS_BRAND  = ["Mk01","Mk02","Mk03","Mk04","Mk05",
                     "Mk06","Mk07","Mk08","Mk09","Mk10"]
INSTANCIAS_DAUZ   = ["01a","04a","07a","10a","13a","16a"]
ORDEN = {"Hurink": INSTANCIAS_HURINK,
         "Brandimarte": INSTANCIAS_BRAND,
         "Dauzere": INSTANCIAS_DAUZ}
COLORS = {"edata": "#4C72B0", "rdata": "#DD8452", "vdata": "#55A868"}


#lectura
def buscar_xlsx(dirs):
    archivos = []
    for d in dirs:
        encontrados = glob.glob(os.path.join(d, "resultados.xlsx"))
        encontrados += glob.glob(os.path.join(d, "**", "resultados.xlsx"),
                                 recursive=True)
        archivos += encontrados
    return sorted(set(archivos))


def leer_hojas(rutas):
    resumen, convergencia = [], []
    tiene_conv = False
    for ruta in rutas:
        wb = load_workbook(ruta, data_only=True)
        # resumen
        if "Resumen_General" in wb.sheetnames:
            ws  = wb["Resumen_General"]
            enc = [c.value for c in ws[1]]
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if fila[0] is None:
                    continue
                resumen.append(dict(zip(enc, fila)))
        # convergencia
        if "Convergencia" in wb.sheetnames:
            tiene_conv = True
            ws  = wb["Convergencia"]
            enc = [c.value for c in ws[1]]
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if fila[0] is None:
                    continue
                convergencia.append(dict(zip(enc, fila)))
    return resumen, convergencia, tiene_conv


#gap%
def grafico_gap_hurink(filas, out):
    datos = {s: {} for s in SUBSETS_HURINK}
    for r in filas:
        if r["Benchmark"] == "Hurink":
            v = r["Gap%"]
            if v != "" and v is not None:
                datos[r["Grupo"]][r["Instancia"]] = float(v)
    x, ancho = range(len(INSTANCIAS_HURINK)), 0.25
    plt.figure(figsize=(9, 5))
    for k, sub in enumerate(SUBSETS_HURINK):
        vals = [datos[sub].get(i, 0) for i in INSTANCIAS_HURINK]
        pos  = [i + (k - 1) * ancho for i in x]
        bars = plt.bar(pos, vals, ancho, label=sub, color=COLORS[sub])
        for b, v in zip(bars, vals):
            plt.text(b.get_x() + b.get_width()/2, b.get_height(),
                     f"{v:.1f}", ha="center", va="bottom", fontsize=7)
    plt.xticks(list(x), INSTANCIAS_HURINK)
    plt.ylabel("gap% respecto al UB")
    plt.title("Gap% - Hurink (por nivel de flexibilidad)")
    plt.legend(title="subset")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out, dpi=140); plt.close()
    print("  guardado:", out)


def grafico_gap_simple(filas, benchmark, out):
    orden = ORDEN[benchmark]
    d = {r["Instancia"]: float(r["Gap%"]) for r in filas
         if r["Benchmark"] == benchmark
         and r["Gap%"] not in ("", None)}
    vals = [d.get(i, 0) for i in orden]
    plt.figure(figsize=(max(7, len(orden)*0.9), 5))
    bars = plt.bar(orden, vals, color="#4C72B0", edgecolor="black")
    for b, v in zip(bars, vals):
        plt.text(b.get_x() + b.get_width()/2, b.get_height(),
                 f"{v:.1f}", ha="center", va="bottom", fontsize=7)
    plt.ylabel("gap% respecto al UB")
    plt.title(f"Gap% - {benchmark}")
    plt.xticks(rotation=45, ha="right")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out, dpi=140); plt.close()
    print("  guardado:", out)


#makespan
def grafico_makespan_hurink(filas, out):
    prom = {s: {} for s in SUBSETS_HURINK}
    desv = {s: {} for s in SUBSETS_HURINK}
    for r in filas:
        if r["Benchmark"] == "Hurink":
            prom[r["Grupo"]][r["Instancia"]] = float(r["Promedio"])
            desv[r["Grupo"]][r["Instancia"]] = float(r["DesvEst"])
    x, ancho = range(len(INSTANCIAS_HURINK)), 0.25
    plt.figure(figsize=(9, 5))
    for k, sub in enumerate(SUBSETS_HURINK):
        ys = [prom[sub].get(i, 0) for i in INSTANCIAS_HURINK]
        es = [desv[sub].get(i, 0) for i in INSTANCIAS_HURINK]
        pos = [i + (k-1)*ancho for i in x]
        plt.bar(pos, ys, ancho, yerr=es, capsize=3,
                label=sub, color=COLORS[sub])
    plt.xticks(list(x), INSTANCIAS_HURINK)
    plt.ylabel("Makespan promedio +/- desv. estandar")
    plt.title("Makespan +/- desviacion - Hurink")
    plt.legend(title="subset")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out, dpi=140); plt.close()
    print("  guardado:", out)


def grafico_makespan_simple(filas, benchmark, out):
    orden = ORDEN[benchmark]
    prom = {r["Instancia"]: float(r["Promedio"]) for r in filas
            if r["Benchmark"] == benchmark}
    desv = {r["Instancia"]: float(r["DesvEst"]) for r in filas
            if r["Benchmark"] == benchmark}
    ys = [prom.get(i, 0) for i in orden]
    es = [desv.get(i, 0) for i in orden]
    plt.figure(figsize=(max(7, len(orden)*0.9), 5))
    plt.bar(orden, ys, yerr=es, capsize=3, color="#937860", edgecolor="black")
    plt.ylabel("Makespan promedio +/- desv. estandar")
    plt.title(f"Makespan +/- desviacion - {benchmark}")
    plt.xticks(rotation=45, ha="right")
    plt.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out, dpi=140); plt.close()
    print("  guardado:", out)


#convergencia
def grafico_convergencia(convergencia, benchmark, group, titulo, out):
    curvas = defaultdict(list)
    for r in convergencia:
        if r["Benchmark"] == benchmark and r["Grupo"] == group:
            curvas[r["Instancia"]].append(
                (int(r["Iteracion"]), float(r["MejorMakespan"])))
    if not curvas:
        print(f"  [sin datos de convergencia para {titulo}]")
        return
    plt.figure(figsize=(8, 5))
    for inst in sorted(curvas.keys()):
        pts = sorted(curvas[inst])
        xs  = [p[0] for p in pts]
        ys  = [p[1] for p in pts]
        plt.plot(xs, ys, drawstyle="steps-post",
                 marker="o", markersize=3, label=inst)
    plt.xlabel("Iteracion")
    plt.ylabel("Mejor makespan (Cmax) hasta el momento")
    plt.title(f"Convergencia del ACO - {titulo}")
    plt.legend(title="instancia", fontsize=8)
    plt.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(out, dpi=140); plt.close()
    print("  guardado:", out)


#main
def main():
    ap = argparse.ArgumentParser(description="Graficos FJSP-ACO (11 figuras)")
    ap.add_argument("--dirs", nargs="+", default=["."],
                    help="carpetas con resultados.xlsx")
    ap.add_argument("--out", default="graficos")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)

    rutas = buscar_xlsx(args.dirs)
    if not rutas:
        print("No encontre ningun resultados.xlsx en:", args.dirs)
        return
    print("Leyendo:", ", ".join(rutas))

    resumen, convergencia, tiene_conv = leer_hojas(rutas)
    benchmarks = sorted({r["Benchmark"] for r in resumen})
    print("Benchmarks:", benchmarks)
    print(f"Filas de convergencia encontradas: {len(convergencia)}")
    if not tiene_conv:
        print("  [AVISO] Ningun Excel tiene hoja Convergencia.")
        print("  Los graficos de convergencia no se generaran.")
        print("  Asegurate de haber corrido con el fjsp_aco.py nuevo.")

    print("\n[1-3] Gap%")
    if "Hurink"       in benchmarks:
        grafico_gap_hurink(resumen,
            os.path.join(args.out, "gap_Hurink.png"))
    if "Brandimarte"  in benchmarks:
        grafico_gap_simple(resumen, "Brandimarte",
            os.path.join(args.out, "gap_Brandimarte.png"))
    if "Dauzere"      in benchmarks:
        grafico_gap_simple(resumen, "Dauzere",
            os.path.join(args.out, "gap_Dauzere.png"))

    print("\n[4-6] Makespan +/- desviacion")
    if "Hurink"       in benchmarks:
        grafico_makespan_hurink(resumen,
            os.path.join(args.out, "makespan_std_Hurink.png"))
    if "Brandimarte"  in benchmarks:
        grafico_makespan_simple(resumen, "Brandimarte",
            os.path.join(args.out, "makespan_std_Brandimarte.png"))
    if "Dauzere"      in benchmarks:
        grafico_makespan_simple(resumen, "Dauzere",
            os.path.join(args.out, "makespan_std_Dauzere.png"))

    if tiene_conv:
        print("\n[7-11] Convergencia (desde la hoja Convergencia del Excel)")
        objetivos = [
            ("Hurink",      "edata",       "Hurink edata"),
            ("Hurink",      "rdata",       "Hurink rdata"),
            ("Hurink",      "vdata",       "Hurink vdata"),
            ("Brandimarte", "Brandimarte", "Brandimarte"),
            ("Dauzere",     "Dauzere",     "Dauzere"),
        ]
        for bm, gr, titulo in objetivos:
            if bm not in benchmarks:
                continue
            nombre = "convergencia_" + titulo.replace(" ", "_") + ".png"
            grafico_convergencia(convergencia, bm, gr, titulo,
                                 os.path.join(args.out, nombre))

    total = len(glob.glob(os.path.join(args.out, "*.png")))
    print(f"\nListo. {total} graficos en la carpeta: {args.out}")


if __name__ == "__main__":
    main()
