import os
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.lines import Line2D
from openpyxl import load_workbook

DIR_BASE     = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH    = os.path.join(DIR_BASE, "resultados", "resultados.xlsx")
DIR_GRAFICOS = os.path.join(DIR_BASE, "graficos")
os.makedirs(DIR_GRAFICOS, exist_ok=True)

COLORES = ["#2E4057", "#048A81", "#E07A5F", "#3D405B", "#81B29A",
           "#F2CC8F", "#9B2226", "#AE2012", "#CA6702", "#EE9B00"]

plt.rcParams.update({
    "font.family":    "DejaVu Sans",
    "font.size":      10,
    "axes.titlesize": 11,
    "axes.labelsize": 10,
    "legend.fontsize": 8,
    "figure.dpi":     150,
})


def leer_hoja_resumen(wb, nombre_hoja):
    ws       = wb[nombre_hoja]
    cabecera = None
    data     = []
    for row in ws.iter_rows(values_only=True):
        if row[0] == "Instancia":
            cabecera = row
            continue
        if cabecera and row[0] and isinstance(row[0], str):
            data.append(dict(zip(cabecera, row)))
    return data


def leer_convergencia(wb, nombre_hoja):
    ws     = wb[nombre_hoja]
    result = {}
    inst   = None
    leyendo = False

    for row in ws.iter_rows(values_only=True):
        # fila de nombre de instancia
        if (isinstance(row[0], str) and row[1] is None
                and row[0] != "Iteracion" and "Conv" not in row[0]):
            inst    = row[0]
            leyendo = False
            result[inst] = {"iters": [], "prom": [], "reps": []}
            continue
        # fila de cabecera
        if row[0] == "Iteracion":
            leyendo = True
            continue
        # fila de datos
        if leyendo and inst and isinstance(row[0], (int, float)):
            result[inst]["iters"].append(row[0])
            result[inst]["prom"].append(row[-1])
            result[inst]["reps"].append(list(row[1:-1]))
            continue
        # fila vacia — fin del bloque
        if leyendo and row[0] is None:
            leyendo = False
            inst    = None

    return result


def guardar(fig, nombre):
    ruta = os.path.join(DIR_GRAFICOS, nombre)
    fig.savefig(ruta, bbox_inches="tight")
    plt.close(fig)
    print(f"  [OK] {nombre}")


# =============================================================================
# GRAFICOS 1-3: GAP%
# =============================================================================

def grafico_gap(wb, hojas, titulo, nombre_archivo, colores_flex=None):
    fig, ax = plt.subplots(figsize=(8, 4))
    x_ticks  = []
    x_labels = []
    x = 0

    for i, (label, hoja) in enumerate(hojas):
        data  = leer_hoja_resumen(wb, hoja)
        insts = [d["Instancia"] for d in data]
        gaps  = [d["Gap%UB"] if d["Gap%UB"] != "" else 0 for d in data]
        color = COLORES[i] if colores_flex is None else colores_flex[i]
        xs    = list(range(x, x + len(insts)))
        ax.bar(xs, gaps, color=color, label=label, width=0.7, zorder=3)
        x_ticks.extend(xs)
        x_labels.extend(insts)
        x += len(insts) + 1

    ax.axhline(0, color="black", linewidth=0.8, linestyle="--", zorder=2)
    ax.set_xticks(x_ticks)
    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Gap% respecto al UB")
    ax.set_title(titulo)
    ax.legend(loc="upper right")
    ax.grid(axis="y", linestyle="--", alpha=0.5, zorder=0)
    ax.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.1f%%"))
    fig.tight_layout()
    guardar(fig, nombre_archivo)


def graficos_gap(wb):
    grafico_gap(wb,
        [("edata (flex. baja)",  "Hurink_edata"),
         ("rdata (flex. media)", "Hurink_rdata"),
         ("vdata (flex. alta)",  "Hurink_vdata")],
        "Gap% vs UB — Hurink (edata / rdata / vdata)",
        "gap_hurink.png",
        colores_flex=["#2E4057", "#048A81", "#E07A5F"])

    grafico_gap(wb,
        [("Brandimarte", "Brandimarte")],
        "Gap% vs UB — Brandimarte",
        "gap_brandimarte.png")

    grafico_gap(wb,
        [("Dauzere", "Dauzere")],
        "Gap% vs UB — Dauzere",
        "gap_dauzere.png")


# =============================================================================
# GRAFICOS 4-6: MAKESPAN +/- STD
# =============================================================================

def grafico_makespan(wb, hojas, titulo, nombre_archivo, colores_flex=None):
    fig, ax = plt.subplots(figsize=(8, 4))
    x_ticks  = []
    x_labels = []
    x = 0

    for i, (label, hoja) in enumerate(hojas):
        data  = leer_hoja_resumen(wb, hoja)
        insts = [d["Instancia"] for d in data]
        proms = [d["Promedio"] for d in data]
        stds  = [d["Std"] for d in data]
        ubs   = [d["UB"] if d["UB"] != "" else None for d in data]
        color = COLORES[i] if colores_flex is None else colores_flex[i]
        xs    = list(range(x, x + len(insts)))
        ax.bar(xs, proms, yerr=stds, color=color, label=label,
               width=0.7, capsize=4, zorder=3,
               error_kw={"elinewidth": 1.2, "ecolor": "black"})
        for xi, ub in zip(xs, ubs):
            if ub:
                ax.hlines(ub, xi - 0.4, xi + 0.4,
                          colors="red", linewidths=1.2,
                          linestyles="--", zorder=4)
        x_ticks.extend(xs)
        x_labels.extend(insts)
        x += len(insts) + 1

    ax.set_xticks(x_ticks)
    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=8)
    ax.set_ylabel("Makespan")
    ax.set_title(titulo)
    handles, labels = ax.get_legend_handles_labels()
    handles.append(Line2D([0], [0], color="red", linewidth=1.2,
                          linestyle="--", label="UB literatura"))
    ax.legend(handles=handles, loc="upper left", fontsize=8)
    ax.grid(axis="y", linestyle="--", alpha=0.5, zorder=0)
    fig.tight_layout()
    guardar(fig, nombre_archivo)


def graficos_makespan(wb):
    grafico_makespan(wb,
        [("edata (flex. baja)",  "Hurink_edata"),
         ("rdata (flex. media)", "Hurink_rdata"),
         ("vdata (flex. alta)",  "Hurink_vdata")],
        "Makespan promedio ± std — Hurink",
        "makespan_hurink.png",
        colores_flex=["#2E4057", "#048A81", "#E07A5F"])

    grafico_makespan(wb,
        [("Brandimarte", "Brandimarte")],
        "Makespan promedio ± std — Brandimarte",
        "makespan_brandimarte.png")

    grafico_makespan(wb,
        [("Dauzere", "Dauzere")],
        "Makespan promedio ± std — Dauzere",
        "makespan_dauzere.png")


# =============================================================================
# GRAFICOS 7-11: CONVERGENCIA
# =============================================================================

def grafico_convergencia(wb, hoja_conv, titulo, nombre_archivo):
    conv = leer_convergencia(wb, hoja_conv)
    if not conv:
        print(f"  [SKIP] sin datos: {nombre_archivo}")
        return

    fig, ax = plt.subplots(figsize=(8, 4))
    for i, (inst, d) in enumerate(conv.items()):
        color = COLORES[i % len(COLORES)]
        ax.plot(d["iters"], d["prom"], color=color,
                linewidth=1.5, label=inst, zorder=3)

    ax.set_xlabel("Iteracion VNS")
    ax.set_ylabel("Makespan")
    ax.set_title(titulo)
    ax.legend(loc="upper right", fontsize=8)
    ax.grid(linestyle="--", alpha=0.4, zorder=0)
    fig.tight_layout()
    guardar(fig, nombre_archivo)


def graficos_convergencia(wb):
    configs = [
        ("Conv_Hurink_edata", "Convergencia — Hurink edata (flex. baja)",  "conv_hurink_edata.png"),
        ("Conv_Hurink_rdata", "Convergencia — Hurink rdata (flex. media)", "conv_hurink_rdata.png"),
        ("Conv_Hurink_vdata", "Convergencia — Hurink vdata (flex. alta)",  "conv_hurink_vdata.png"),
        ("Conv_Brandimarte",  "Convergencia — Brandimarte",                "conv_brandimarte.png"),
        ("Conv_Dauzere",      "Convergencia — Dauzere",                    "conv_dauzere.png"),
    ]
    for hoja, titulo, archivo in configs:
        grafico_convergencia(wb, hoja, titulo, archivo)


# =============================================================================
# MAIN
# =============================================================================

def main():
    print(f"Leyendo {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH, data_only=True)

    print("\nGenerando graficos Gap%...")
    graficos_gap(wb)

    print("\nGenerando graficos Makespan...")
    graficos_makespan(wb)

    print("\nGenerando graficos Convergencia...")
    graficos_convergencia(wb)

    total = len(os.listdir(DIR_GRAFICOS))
    print(f"\nListo. {total} graficos en: {DIR_GRAFICOS}")


if __name__ == "__main__":
    main()
