import os
import time
import datetime
import statistics
import platform
import sys

from fjsp_core import FJSPInstance
from ts_vns    import ts_vns

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_OK = True
except ImportError:
    XLSX_OK = False
    print("[AVISO] openpyxl no instalado. Solo se generara el .txt.")
    print("        Instalar con: pip install openpyxl")


# =============================================================================
# CONFIGURACION
# =============================================================================

REPETICIONES = 5
SEMILLA_BASE = 100

PARAMETROS = dict(
    max_iter      = 40,
    k_max         = 4,
    tabu_iter     = 40,
    tenencia      = 10,
    tiempo_limite = 60,
)

HURINK = {
    "edata": ["la01", "la06", "la11", "la16", "la21"],
    "rdata": ["la01", "la06", "la11", "la16", "la21"],
    "vdata": ["la01", "la06", "la11", "la16", "la21"],
}

BRANDIMARTE = ["Mk01", "Mk02", "Mk03", "Mk04", "Mk05",
               "Mk06", "Mk07", "Mk08", "Mk09", "Mk10"]

DAUZERE = ["01a", "04a", "07a", "10a", "13a", "16a"]

UB_HURINK = {
    ("edata","la01"): 609,  ("edata","la06"): 800,  ("edata","la11"): 1071,
    ("edata","la16"): 717,  ("edata","la21"): 835,
    ("rdata","la01"): 570,  ("rdata","la06"): 799,  ("rdata","la11"): 1071,
    ("rdata","la16"): 717,  ("rdata","la21"): 833,
    ("vdata","la01"): 570,  ("vdata","la06"): 799,  ("vdata","la11"): 1071,
    ("vdata","la16"): 717,  ("vdata","la21"): 800,
}

UB_BRANDIMARTE = {
    "Mk01": 42,  "Mk02": 32,  "Mk03": 211, "Mk04": 81,  "Mk05": 186,
    "Mk06": 86,  "Mk07": 157, "Mk08": 523, "Mk09": 369, "Mk10": 296,
}

UB_DAUZERE = {
    "01a": 2530, "04a": 2565, "07a": 2408,
    "10a": 2362, "13a": 2302, "16a": 2301,
}

FLEX_HURINK = {"edata": "Baja", "rdata": "Media", "vdata": "Alta"}


# =============================================================================
# AUXILIARES
# =============================================================================

def detectar_hardware():
    return {
        "procesador": platform.processor() or platform.machine(),
        "sistema":    platform.system() + " " + platform.release(),
        "python":     f"Python {sys.version.split()[0]}",
    }


def pct(valor, ref):
    if ref is None or ref == 0:
        return None
    return round(100.0 * (valor - ref) / ref, 2)


def correr_instancia(inst, nombre, ub):
    resultados  = []
    historicos  = []   # convergencia por repeticion
    for r in range(REPETICIONES):
        semilla = SEMILLA_BASE + r
        t0      = time.time()
        sol, hist = ts_vns(inst, semilla=semilla,
                           registrar_convergencia=True, **PARAMETROS)
        t_total = time.time() - t0
        resultados.append({
            "rep":            r + 1,
            "semilla":        semilla,
            "makespan":       sol.makespan,
            "iter_vns":       getattr(sol, "iter_vns", 0),
            "iter_tabu":      getattr(sol, "iter_tabu", 0),
            "iter_mejor":     getattr(sol, "iter_mejor", 0),
            "tiempo_s":       round(t_total, 3),
            "tiempo_mejor_s": round(getattr(sol, "tiempo_mejor_s", 0.0), 3),
        })
        historicos.append(hist)

    mk     = [r["makespan"] for r in resultados]
    mejor  = min(mk)
    peor   = max(mk)
    prom   = statistics.mean(mk)
    std    = statistics.stdev(mk) if len(mk) > 1 else 0.0
    t_prom = statistics.mean([r["tiempo_s"] for r in resultados])
    gap    = pct(mejor, ub)
    veces  = sum(1 for m in mk if ub and m <= ub)

    resumen = {
        "instancia":  nombre,
        "n_jobs":     inst.n_jobs,
        "n_machines": inst.n_machines,
        "mejor":      mejor,
        "promedio":   round(prom, 1),
        "std":        round(std, 1),
        "peor":       peor,
        "ub":         ub if ub else "",
        "gap_ub":     gap if gap is not None else "",
        "veces_ub":   f"{veces}/{REPETICIONES}",
        "tiempo_s":   round(t_prom, 2),
    }
    return resumen, resultados, historicos


def _conv_promedio(historicos):
    """Promedia los historicos de convergencia alineando por longitud minima."""
    if not historicos:
        return []
    n = min(len(h) for h in historicos)
    return [round(statistics.mean(h[i] for h in historicos), 2) for i in range(n)]


# =============================================================================
# EXPERIMENTO PRINCIPAL
# =============================================================================

def main():
    dir_base   = os.path.dirname(os.path.abspath(__file__))
    dir_inst   = os.path.join(dir_base, "instancias")
    dir_result = os.path.join(dir_base, "resultados")
    os.makedirs(dir_result, exist_ok=True)
    ahora      = datetime.datetime.now()
    hardware   = detectar_hardware()

    datos = {
        "hurink_edata": {"resumenes": [], "reps": [], "conv": []},
        "hurink_rdata": {"resumenes": [], "reps": [], "conv": []},
        "hurink_vdata": {"resumenes": [], "reps": [], "conv": []},
        "brandimarte":  {"resumenes": [], "reps": [], "conv": []},
        "dauzere":      {"resumenes": [], "reps": [], "conv": []},
    }

    SEP  = "-" * 80
    SEP2 = "=" * 80

    print(SEP2)
    print(f"  TS+VNS / FJSP  [{ahora.strftime('%Y-%m-%d %H:%M')}]")
    print(SEP2)

    # --- Hurink ---
    for flex, nombres in HURINK.items():
        print(f"\n  Hurink {flex}")
        print(SEP)
        for nombre in nombres:
            ruta = os.path.join(dir_inst, "Hurink", flex, nombre + ".txt")
            if not os.path.isfile(ruta):
                print(f"  {nombre}: archivo no encontrado")
                continue
            inst = FJSPInstance(ruta, nombre, flex)
            ub   = UB_HURINK.get((flex, nombre))
            res, reps, hists = correr_instancia(inst, nombre, ub)
            res["benchmark"]  = "Hurink"
            res["subfamilia"] = flex
            for r in reps:
                r["benchmark"]  = "Hurink"
                r["subfamilia"] = flex
                r["instancia"]  = nombre
            clave = f"hurink_{flex}"
            datos[clave]["resumenes"].append(res)
            datos[clave]["reps"].extend(reps)
            datos[clave]["conv"].append({
                "instancia": nombre,
                "historicos": hists,
                "promedio":   _conv_promedio(hists),
            })
            gap_s = f"{res['gap_ub']:.2f}%" if res['gap_ub'] != "" else "—"
            print(f"  {nombre:<8} {inst.n_jobs}x{inst.n_machines}  "
                  f"mejor={res['mejor']}  prom={res['promedio']}  gap={gap_s}")

    # --- Brandimarte ---
    print(f"\n  Brandimarte")
    print(SEP)
    for nombre in BRANDIMARTE:
        ruta = os.path.join(dir_inst, "Brandimarte", nombre + ".fjs")
        if not os.path.isfile(ruta):
            print(f"  {nombre}: archivo no encontrado")
            continue
        inst = FJSPInstance(ruta, nombre, "Brandimarte")
        ub   = UB_BRANDIMARTE.get(nombre)
        res, reps, hists = correr_instancia(inst, nombre, ub)
        res["benchmark"]  = "Brandimarte"
        res["subfamilia"] = ""
        for r in reps:
            r["benchmark"]  = "Brandimarte"
            r["subfamilia"] = ""
            r["instancia"]  = nombre
        datos["brandimarte"]["resumenes"].append(res)
        datos["brandimarte"]["reps"].extend(reps)
        datos["brandimarte"]["conv"].append({
            "instancia": nombre,
            "historicos": hists,
            "promedio":   _conv_promedio(hists),
        })
        gap_s = f"{res['gap_ub']:.2f}%" if res['gap_ub'] != "" else "—"
        print(f"  {nombre:<8} {inst.n_jobs}x{inst.n_machines}  "
              f"mejor={res['mejor']}  prom={res['promedio']}  gap={gap_s}")

    # --- Dauzere ---
    print(f"\n  Dauzere")
    print(SEP)
    for nombre in DAUZERE:
        ruta = os.path.join(dir_inst, "Dauzere", nombre + ".fjs")
        if not os.path.isfile(ruta):
            print(f"  {nombre}: archivo no encontrado")
            continue
        inst = FJSPInstance(ruta, nombre, "Dauzere")
        ub   = UB_DAUZERE.get(nombre)
        res, reps, hists = correr_instancia(inst, nombre, ub)
        res["benchmark"]  = "Dauzere"
        res["subfamilia"] = ""
        for r in reps:
            r["benchmark"]  = "Dauzere"
            r["subfamilia"] = ""
            r["instancia"]  = nombre
        datos["dauzere"]["resumenes"].append(res)
        datos["dauzere"]["reps"].extend(reps)
        datos["dauzere"]["conv"].append({
            "instancia": nombre,
            "historicos": hists,
            "promedio":   _conv_promedio(hists),
        })
        gap_s = f"{res['gap_ub']:.2f}%" if res['gap_ub'] != "" else "—"
        print(f"  {nombre:<8} {inst.n_jobs}x{inst.n_machines}  "
              f"mejor={res['mejor']}  prom={res['promedio']}  gap={gap_s}")

    print(f"\n{SEP2}")

    _guardar_reporte_txt(dir_result, datos, ahora, hardware)
    _guardar_configuracion(dir_result, ahora, hardware)
    if XLSX_OK:
        _guardar_xlsx(dir_result, datos)
    else:
        print("[AVISO] xlsx no generado. Instalar openpyxl.")


# =============================================================================
# EXPORTACION TXT
# =============================================================================

def _guardar_reporte_txt(dir_result, datos, ahora, hardware):
    ruta = os.path.join(dir_result, "resultados_reporte.txt")
    SEP  = "-" * 90
    SEP2 = "=" * 90

    CABECERA = (f"  {'Instancia':<10}{'Jobs':>5}{'Maq':>5}"
                f"{'Mejor':>8}{'Prom':>8}{'Std':>7}{'Peor':>7}"
                f"{'UB':>7}{'Gap%':>7}{'UB/#':>7}{'T(s)':>7}")

    with open(ruta, "w", encoding="utf-8") as f:
        f.write(SEP2 + "\n")
        f.write("  REPORTE EXPERIMENTAL — TS+VNS / FJSP\n")
        f.write(f"  Fecha : {ahora.strftime('%Y-%m-%d  %H:%M:%S')}\n")
        f.write(SEP2 + "\n\n")
        f.write("  ENTORNO\n")
        f.write(SEP + "\n")
        for k, v in hardware.items():
            f.write(f"  {k:<14}: {v}\n")
        f.write("\n")

        for flex in ["edata", "rdata", "vdata"]:
            clave = f"hurink_{flex}"
            filas = datos[clave]["resumenes"]
            if not filas:
                continue
            f.write(f"\n  HURINK — {flex.upper()}  (flexibilidad {FLEX_HURINK[flex]})\n")
            f.write(SEP + "\n")
            f.write(CABECERA + "\n")
            f.write(SEP + "\n")
            for r in filas:
                ub_s  = str(r["ub"]) if r["ub"] != "" else "—"
                gap_s = f"{r['gap_ub']:.2f}" if r["gap_ub"] != "" else "—"
                f.write(f"  {r['instancia']:<10}{r['n_jobs']:>5}{r['n_machines']:>5}"
                        f"{r['mejor']:>8}{r['promedio']:>8.1f}{r['std']:>7.1f}{r['peor']:>7}"
                        f"{ub_s:>7}{gap_s:>7}{r['veces_ub']:>7}{r['tiempo_s']:>7.1f}\n")
            gaps = [r["gap_ub"] for r in filas if r["gap_ub"] != ""]
            if gaps:
                f.write(SEP + "\n")
                f.write(f"  Gap% promedio: {statistics.mean(gaps):.2f}%\n")

        for bench, clave in [("BRANDIMARTE", "brandimarte"), ("DAUZERE", "dauzere")]:
            filas = datos[clave]["resumenes"]
            if not filas:
                continue
            f.write(f"\n  {bench}\n")
            f.write(SEP + "\n")
            f.write(CABECERA + "\n")
            f.write(SEP + "\n")
            for r in filas:
                ub_s  = str(r["ub"]) if r["ub"] != "" else "—"
                gap_s = f"{r['gap_ub']:.2f}" if r["gap_ub"] != "" else "—"
                f.write(f"  {r['instancia']:<10}{r['n_jobs']:>5}{r['n_machines']:>5}"
                        f"{r['mejor']:>8}{r['promedio']:>8.1f}{r['std']:>7.1f}{r['peor']:>7}"
                        f"{ub_s:>7}{gap_s:>7}{r['veces_ub']:>7}{r['tiempo_s']:>7.1f}\n")
            gaps = [r["gap_ub"] for r in filas if r["gap_ub"] != ""]
            if gaps:
                f.write(SEP + "\n")
                f.write(f"  Gap% promedio: {statistics.mean(gaps):.2f}%\n")

        f.write("\n\n  LEYENDA\n")
        f.write(SEP + "\n")
        for col, desc in [
            ("Mejor",  "Menor makespan de las 5 repeticiones"),
            ("Prom",   "Makespan promedio"),
            ("Std",    "Desviacion estandar"),
            ("Peor",   "Mayor makespan"),
            ("UB",     "Upper bound de la literatura"),
            ("Gap%",   "100 x (Mejor - UB) / UB"),
            ("UB/#",   "Repeticiones que igualaron el UB"),
            ("T(s)",   "Tiempo promedio de ejecucion en segundos"),
        ]:
            f.write(f"  {col:<8}: {desc}\n")
        f.write("\n" + SEP2 + "\n")

    print(f"  [OK] resultados_reporte.txt -> {ruta}")


# =============================================================================
# EXPORTACION XLSX
# =============================================================================

def _guardar_xlsx(dir_result, datos):
    ruta = os.path.join(dir_result, "resultados.xlsx")
    wb   = Workbook()
    wb.remove(wb.active)

    HEADER_FILL  = PatternFill("solid", fgColor="2E4057")
    HEADER_FONT  = Font(color="FFFFFF", bold=True)
    HEADER_ALIGN = Alignment(horizontal="center")

    COLS_RESUMEN = [
        "Instancia", "Jobs", "Maquinas", "Mejor", "Promedio",
        "Std", "Peor", "UB", "Gap%UB", "Veces_UB", "Tiempo_s"
    ]
    COLS_REPS = [
        "Benchmark", "Subfamilia", "Instancia", "Rep", "Semilla",
        "Makespan", "Iter_VNS", "Iter_Tabu", "Iter_Mejor",
        "Tiempo_s", "Tiempo_Mejor_s"
    ]

    def fila_resumen(r):
        return [
            r["instancia"], r["n_jobs"], r["n_machines"],
            r["mejor"], r["promedio"], r["std"], r["peor"],
            r["ub"] if r["ub"] != "" else "",
            r["gap_ub"] if r["gap_ub"] != "" else "",
            r["veces_ub"], r["tiempo_s"],
        ]

    def fila_rep(r):
        return [
            r["benchmark"], r.get("subfamilia", ""), r["instancia"],
            r["rep"], r["semilla"], r["makespan"],
            r["iter_vns"], r["iter_tabu"], r["iter_mejor"],
            r["tiempo_s"], r["tiempo_mejor_s"],
        ]

    def escribir_hoja(ws, cols, filas_data, titulo=None):
        if titulo:
            ws.append([titulo])
            ws.cell(1, 1).font = Font(bold=True, size=12)
            ws.append([])
        ws.append(cols)
        fila_h = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(fila_h, c)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = HEADER_ALIGN
        for fila in filas_data:
            ws.append(fila)
        for col in ws.columns:
            max_w = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_w + 4

    def escribir_hoja_conv(ws, conv_data, titulo=None):
        """
        Escribe datos de convergencia en formato tabular.
        Columnas: Instancia | Rep_1 | Rep_2 | ... | Rep_N | Promedio
        Cada fila es una iteracion VNS.
        """
        if titulo:
            ws.append([titulo])
            ws.cell(1, 1).font = Font(bold=True, size=12)
            ws.append([])

        for entrada in conv_data:
            nombre   = entrada["instancia"]
            hists    = entrada["historicos"]
            promedio = entrada["promedio"]
            n_iters  = len(promedio)

            # cabecera de la instancia
            cols_conv = (["Iteracion"] +
                         [f"Rep_{r+1}" for r in range(len(hists))] +
                         ["Promedio"])
            ws.append([nombre])
            ws.cell(ws.max_row, 1).font = Font(bold=True)
            ws.append(cols_conv)
            fila_h = ws.max_row
            for c in range(1, len(cols_conv) + 1):
                cell = ws.cell(fila_h, c)
                cell.fill      = HEADER_FILL
                cell.font      = HEADER_FONT
                cell.alignment = HEADER_ALIGN

            for i in range(n_iters):
                fila = [i + 1]
                for h in hists:
                    fila.append(h[i] if i < len(h) else "")
                fila.append(promedio[i])
                ws.append(fila)
            ws.append([])

        for col in ws.columns:
            max_w = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_w + 3

    # Resumen general
    ws    = wb.create_sheet("Resumen_General")
    todas = []
    for clave in ["hurink_edata","hurink_rdata","hurink_vdata","brandimarte","dauzere"]:
        for r in datos[clave]["resumenes"]:
            label = f"{r.get('benchmark','')} {r.get('subfamilia','')}".strip()
            todas.append([label] + fila_resumen(r))
    escribir_hoja(ws, ["Benchmark"] + COLS_RESUMEN, todas,
                  "Resumen General — TS+VNS / FJSP")

    # Hurink por flexibilidad
    for flex in ["edata", "rdata", "vdata"]:
        ws    = wb.create_sheet(f"Hurink_{flex}")
        filas = [fila_resumen(r) for r in datos[f"hurink_{flex}"]["resumenes"]]
        escribir_hoja(ws, COLS_RESUMEN, filas,
                      f"Hurink {flex} — flexibilidad {FLEX_HURINK[flex]}")

    # Brandimarte y Dauzere
    for nombre_hoja, clave in [("Brandimarte","brandimarte"), ("Dauzere","dauzere")]:
        ws    = wb.create_sheet(nombre_hoja)
        filas = [fila_resumen(r) for r in datos[clave]["resumenes"]]
        escribir_hoja(ws, COLS_RESUMEN, filas, nombre_hoja)

    # Repeticiones por benchmark
    for nombre_hoja, clave in [
        ("Rep_Hurink_edata", "hurink_edata"),
        ("Rep_Hurink_rdata", "hurink_rdata"),
        ("Rep_Hurink_vdata", "hurink_vdata"),
        ("Rep_Brandimarte",  "brandimarte"),
        ("Rep_Dauzere",      "dauzere"),
    ]:
        ws    = wb.create_sheet(nombre_hoja)
        filas = [fila_rep(r) for r in datos[clave]["reps"]]
        escribir_hoja(ws, COLS_REPS, filas, nombre_hoja.replace("_", " "))

    # Convergencia por benchmark
    for nombre_hoja, clave in [
        ("Conv_Hurink_edata", "hurink_edata"),
        ("Conv_Hurink_rdata", "hurink_rdata"),
        ("Conv_Hurink_vdata", "hurink_vdata"),
        ("Conv_Brandimarte",  "brandimarte"),
        ("Conv_Dauzere",      "dauzere"),
    ]:
        ws = wb.create_sheet(nombre_hoja)
        escribir_hoja_conv(ws, datos[clave]["conv"],
                           nombre_hoja.replace("_", " "))

    wb.save(ruta)
    print(f"  [OK] resultados.xlsx -> {ruta}")


# =============================================================================
# CONFIGURACION
# =============================================================================

def _guardar_configuracion(dir_result, ahora, hardware):
    ruta = os.path.join(dir_result, "configuracion.txt")
    with open(ruta, "w", encoding="utf-8") as f:
        f.write("=" * 60 + "\n")
        f.write("  CONFIGURACION DEL EXPERIMENTO\n")
        f.write(f"  {ahora.strftime('%Y-%m-%d  %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")
        f.write("  Algoritmo    : TS+VNS con agitacion adaptativa\n")
        f.write(f"  Repeticiones : {REPETICIONES}\n")
        f.write(f"  Semilla base : {SEMILLA_BASE}  "
                f"(semillas {SEMILLA_BASE}..{SEMILLA_BASE+REPETICIONES-1})\n\n")
        f.write("  Parametros:\n")
        for k, v in PARAMETROS.items():
            f.write(f"    {k:<20}: {v}\n")
        f.write("\n  Instancias:\n")
        f.write(f"    Hurink      : {sum(len(v) for v in HURINK.values())} "
                f"(edata/rdata/vdata x {len(HURINK['edata'])})\n")
        f.write(f"    Brandimarte : {len(BRANDIMARTE)}\n")
        f.write(f"    Dauzere     : {len(DAUZERE)}\n")
        f.write("\n  Entorno:\n")
        for k, v in hardware.items():
            f.write(f"    {k:<14}: {v}\n")
        f.write("\n" + "=" * 60 + "\n")
    print(f"  [OK] configuracion.txt -> {ruta}")


if __name__ == "__main__":
    main()
