"""
experimentos.py
===============
Runner de experimentos del Algoritmo Memetico para el FJSP sobre tres
benchmarks:

  - Hurink (1994)     : subconjuntos edata / rdata / vdata, instancias
                        la01, la06, la11, la16, la21 (formato .txt).
  - Brandimarte (1993): instancias Mk01..Mk10 (formato .fjs).
  - Dauzere (1998)    : instancias 01a, 04a, 07a, 10a, 13a, 16a (formato .fjs).
"""

import os
import time
import math
import platform
import argparse
from datetime import datetime

from instancia import leer_instancia
from memetico import memetico

# DEFINICION DE BENCHMARKS Y UPPER BOUNDS DE REFERENCIA

# Upper bounds de la literatura. Para Brandimarte y Dauzere se usan los
# valores de Mastrolilli & Gambardella (2000); para Hurink, los mejores
# valores conocidos del benchmark original.

UB_HURINK = {
    "edata": {"la01": 609, "la06": 800, "la11": 1071, "la16": 717, "la21": 835},
    "rdata": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 833},
    "vdata": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 800},
}

UB_BRANDIMARTE = {
    "Mk01": 42, "Mk02": 32, "Mk03": 211, "Mk04": 81, "Mk05": 186,
    "Mk06": 86, "Mk07": 157, "Mk08": 523, "Mk09": 369, "Mk10": 296,
}

UB_DAUZERE = {
    "01a": 2530, "04a": 2565, "07a": 2408,
    "10a": 2362, "13a": 2302, "16a": 2301,
}

# Subconjuntos e instancias por benchmark.
HURINK_SUBCONJUNTOS = ["edata", "rdata", "vdata"]
HURINK_INSTANCIAS = ["la01", "la06", "la11", "la16", "la21"]
BRANDIMARTE_INSTANCIAS = ["Mk01", "Mk02", "Mk03", "Mk04", "Mk05",
                          "Mk06", "Mk07", "Mk08", "Mk09", "Mk10"]
DAUZERE_INSTANCIAS = ["01a", "04a", "07a", "10a", "13a", "16a"]

# UTILIDADES ESTADISTICAS 

def media(xs):
    """Media aritmetica de una lista de numeros."""
    return sum(xs) / len(xs)


def desviacion_estandar_muestral(xs):
    """Desviacion estandar muestral (ddof=1).

    Devuelve 0.0 si hay una sola observacion.
    """
    n = len(xs)
    if n < 2:
        return 0.0
    m = media(xs)
    var = sum((x - m) ** 2 for x in xs) / (n - 1)
    return math.sqrt(var)

# EJECUCION DE UNA INSTANCIA

def ejecutar_instancia(ruta, etiqueta, ub, tiempo_max, n_reps, semilla_base):
    """Ejecuta n_reps repeticiones del memetico sobre una instancia.

    Devuelve una tupla (fila_resumen, filas_repeticion) donde fila_resumen es
    un diccionario con las metricas agregadas y filas_repeticion es una lista
    de diccionarios, uno por repeticion.
    """
    inst = leer_instancia(ruta, etiqueta)

    makespans, iters, tiempos = [], [], []
    filas_rep = []

    for r in range(n_reps):
        semilla = semilla_base + r
        res = memetico(inst, semilla=semilla, tiempo_max=tiempo_max)
        mk = res["makespan"]
        gap = 100.0 * (mk - ub) / ub

        makespans.append(mk)
        iters.append(res["iteraciones"])
        tiempos.append(res["tiempo"])
        filas_rep.append({
            "instancia": etiqueta,
            "repeticion": r + 1,
            "semilla": semilla,
            "makespan": mk,
            "gap_pct": gap,
            "iteraciones": res["iteraciones"],
            "evaluaciones": res["evaluaciones"],
            "tiempo_s": res["tiempo"],
        })

    mejor = min(makespans)
    prom = media(makespans)
    desv = desviacion_estandar_muestral(makespans)
    peor = max(makespans)

    fila_resumen = {
        "instancia": etiqueta,
        "n_trabajos": inst.n_trabajos,
        "n_maquinas": inst.n_maquinas,
        "total_ops": inst.total_ops,
        "ub": ub,
        "mejor": mejor,
        "promedio": prom,
        "desv_std": desv,
        "peor": peor,
        "gap_mejor_pct": 100.0 * (mejor - ub) / ub,
        "gap_prom_pct": 100.0 * (prom - ub) / ub,
        "iter_promedio": media(iters),
        "tiempo_promedio": media(tiempos),
    }
    return fila_resumen, filas_rep


# RECORRIDO COMPLETO DE LOS TRES BENCHMARKS

def correr_todo(instancias_dir, tiempo_max, n_reps, semilla_base, registro=print):
    # Recorre Hurink, Brandimarte y Dauzere.

    grupos = {}
    general = []

    def procesar(nombre_grupo, benchmark, ruta, etiqueta, ub):
        registro("\n[%s]  UB=%s" % (etiqueta, ub))
        fila, filas_rep = ejecutar_instancia(
            ruta, etiqueta, ub, tiempo_max, n_reps, semilla_base)
        grupos.setdefault(nombre_grupo, {"resumen": [], "rep": []})
        grupos[nombre_grupo]["resumen"].append(fila)
        grupos[nombre_grupo]["rep"].extend(filas_rep)
        g = dict(fila)
        g["benchmark"] = benchmark
        g["grupo"] = nombre_grupo
        general.append(g)
        registro("  -> mejor=%d  prom=%.1f  desv=%.2f  peor=%d  gapMejor=%.2f%%"
                 % (fila["mejor"], fila["promedio"], fila["desv_std"],
                    fila["peor"], fila["gap_mejor_pct"]))
        for fr in filas_rep:
            registro("    rep %d (semilla %d): makespan=%d  gap=%.2f%%  gen=%d  t=%.1fs"
                     % (fr["repeticion"], fr["semilla"], fr["makespan"],
                        fr["gap_pct"], fr["iteraciones"], fr["tiempo_s"]))

    #  Hurink 
    registro("\n" + "=" * 70)
    registro("BENCHMARK: HURINK (1994)")
    registro("=" * 70)
    for sub in HURINK_SUBCONJUNTOS:
        for inst in HURINK_INSTANCIAS:
            ruta = os.path.join(instancias_dir, "Hurink", sub, inst + ".txt")
            ub = UB_HURINK[sub][inst]
            procesar("Hurink_" + sub, "Hurink", ruta, "%s/%s" % (sub, inst), ub)

    # Brandimarte 
    registro("\n" + "=" * 70)
    registro("BENCHMARK: BRANDIMARTE (1993)")
    registro("=" * 70)
    for inst in BRANDIMARTE_INSTANCIAS:
        ruta = os.path.join(instancias_dir, "Brandimarte", inst + ".fjs")
        ub = UB_BRANDIMARTE[inst]
        procesar("Brandimarte", "Brandimarte", ruta, inst, ub)

    #  Dauzere 
    registro("\n" + "=" * 70)
    registro("BENCHMARK: DAUZERE (1998)")
    registro("=" * 70)
    for inst in DAUZERE_INSTANCIAS:
        ruta = os.path.join(instancias_dir, "Dauzere", inst + ".fjs")
        ub = UB_DAUZERE[inst]
        procesar("Dauzere", "Dauzere", ruta, inst, ub)

    return {"grupos": grupos, "general": general}


# EXPORTACION: REPORTE DE TEXTO

def _bloque_tabla(filas):
    """Construye las lineas de una tabla de resumen para el reporte TXT."""
    cab = ("%-12s %4s %4s %5s %7s %7s %9s %8s %7s %9s %9s %8s %9s" %
           ("instancia", "nT", "nM", "ops", "UB", "mejor", "promedio",
            "desvStd", "peor", "gapMej%", "gapProm%", "iterPr", "tiempoPr"))
    sep = "-" * len(cab)
    out = [cab, sep]
    for f in filas:
        out.append(
            "%-12s %4d %4d %5d %7d %7d %9.2f %8.2f %7d %9.2f %9.2f %8.1f %9.2f"
            % (f["instancia"], f["n_trabajos"], f["n_maquinas"], f["total_ops"],
               f["ub"], f["mejor"], f["promedio"], f["desv_std"], f["peor"],
               f["gap_mejor_pct"], f["gap_prom_pct"], f["iter_promedio"],
               f["tiempo_promedio"]))
    return out


def exportar_reporte_txt(ruta_txt, datos, config):
    """Genera resultados_reporte.txt organizado por benchmark."""
    grupos = datos["grupos"]
    lineas = []
    sep = "=" * 100
    lineas.append(sep)
    lineas.append("INFORME DE EXPERIMENTOS - ALGORITMO MEMETICO PARA FJSP")
    lineas.append("Benchmarks: Hurink (1994) / Brandimarte (1993) / Dauzere (1998)")
    lineas.append(sep)
    lineas.append("")
    lineas.append("Fecha de ejecucion  : %s" % config["timestamp_legible"])
    lineas.append("Metaheuristica      : Algoritmo Memetico (GA + Busqueda Local)")
    lineas.append("Presupuesto/ejec.   : %d s" % config["tiempo_max"])
    lineas.append("Repeticiones/inst.  : %d" % config["n_reps"])
    lineas.append("Semillas            : %d..%d (base %d)"
                  % (config["semilla_base"],
                     config["semilla_base"] + config["n_reps"] - 1,
                     config["semilla_base"]))
    lineas.append("Total de ejecuciones: %d" % config["total_ejecuciones"])
    lineas.append("")

    orden_grupos = [
        ("HURINK - edata", "Hurink_edata"),
        ("HURINK - rdata", "Hurink_rdata"),
        ("HURINK - vdata", "Hurink_vdata"),
        ("BRANDIMARTE", "Brandimarte"),
        ("DAUZERE", "Dauzere"),
    ]

    for titulo, clave in orden_grupos:
        if clave not in grupos:
            continue
        filas = grupos[clave]["resumen"]
        lineas.append(sep)
        lineas.append("BENCHMARK: %s" % titulo)
        lineas.append(sep)
        lineas.extend(_bloque_tabla(filas))
        gaps = [f["gap_mejor_pct"] for f in filas]
        lineas.append("")
        lineas.append("  Gap%% promedio del MEJOR makespan: %.2f%%" % media(gaps))
        lineas.append("")

    # Resumen global por benchmark.
    lineas.append(sep)
    lineas.append("RESUMEN GLOBAL (gap% promedio del mejor makespan)")
    lineas.append(sep)
    por_benchmark = {}
    for g in datos["general"]:
        por_benchmark.setdefault(g["benchmark"], []).append(g["gap_mejor_pct"])
    for bench in ["Hurink", "Brandimarte", "Dauzere"]:
        if bench in por_benchmark:
            lineas.append("  %-12s : %.2f%%" % (bench, media(por_benchmark[bench])))
    todos = [g["gap_mejor_pct"] for g in datos["general"]]
    lineas.append("  %-12s : %.2f%%" % ("GLOBAL", media(todos)))
    lineas.append("")
    lineas.append("Notas:")
    lineas.append("  - gapMej%  = 100*(mejor - UB)/UB.")
    lineas.append("  - gapProm% = 100*(promedio - UB)/UB.")
    lineas.append("  - desvStd  = desviacion estandar muestral del makespan (ddof=1).")
    lineas.append("  - UB Brandimarte/Dauzere: Mastrolilli & Gambardella (2000).")
    lineas.append("")

    with open(ruta_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas))


# EXPORTACION: LIBRO EXCEL

COLS_RESUMEN = [
    ("instancia", "instancia"), ("n_trabajos", "nT"), ("n_maquinas", "nM"),
    ("total_ops", "ops"), ("ub", "UB"), ("mejor", "mejor"),
    ("promedio", "promedio"), ("desv_std", "desv_std"), ("peor", "peor"),
    ("gap_mejor_pct", "gap_mejor_%"), ("gap_prom_pct", "gap_prom_%"),
    ("iter_promedio", "iter_prom"), ("tiempo_promedio", "tiempo_prom_s"),
]

COLS_GENERAL = [("benchmark", "benchmark"), ("grupo", "grupo")] + COLS_RESUMEN

COLS_REP = [
    ("instancia", "instancia"), ("repeticion", "repeticion"),
    ("semilla", "semilla"), ("makespan", "makespan"),
    ("gap_pct", "gap_%"), ("iteraciones", "iteraciones"),
    ("evaluaciones", "evaluaciones"), ("tiempo_s", "tiempo_s"),
]


def _escribir_hoja(ws, columnas, filas):
    """Escribe una hoja con cabecera en negrita y columnas autoajustadas."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cab_fill = PatternFill("solid", fgColor="2C3E50")
    cab_font = Font(bold=True, color="FFFFFF")
    centrado = Alignment(horizontal="center")

    encabezados = [etq for (_clave, etq) in columnas]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.fill = cab_fill
        celda.font = cab_font
        celda.alignment = centrado

    for fila in filas:
        valores = []
        for clave, _etq in columnas:
            v = fila.get(clave, "")
            if isinstance(v, float):
                v = round(v, 2)
            valores.append(v)
        ws.append(valores)

    # Ancho de columnas segun contenido.
    for i, (_clave, etq) in enumerate(columnas, start=1):
        largo = len(str(etq))
        for fila in filas:
            largo = max(largo, len(str(fila.get(_clave, ""))))
        ws.column_dimensions[get_column_letter(i)].width = largo + 3

    ws.freeze_panes = "A2"


def exportar_xlsx(ruta_xlsx, datos):
    """Genera resultados.xlsx con todas las hojas requeridas."""
    from openpyxl import Workbook

    grupos = datos["grupos"]
    wb = Workbook()

    # Hoja 1: Resumen_General.
    ws = wb.active
    ws.title = "Resumen_General"
    _escribir_hoja(ws, COLS_GENERAL, datos["general"])

    # Hojas de resumen por grupo.
    resumenes = [
        ("Hurink_edata", "Hurink_edata"),
        ("Hurink_rdata", "Hurink_rdata"),
        ("Hurink_vdata", "Hurink_vdata"),
        ("Brandimarte", "Brandimarte"),
        ("Dauzere", "Dauzere"),
    ]
    for titulo_hoja, clave in resumenes:
        ws = wb.create_sheet(titulo_hoja)
        filas = grupos.get(clave, {}).get("resumen", [])
        _escribir_hoja(ws, COLS_RESUMEN, filas)

    # Hojas de repeticiones por grupo.
    repeticiones = [
        ("Rep_Hurink_edata", "Hurink_edata"),
        ("Rep_Hurink_rdata", "Hurink_rdata"),
        ("Rep_Hurink_vdata", "Hurink_vdata"),
        ("Rep_Brandimarte", "Brandimarte"),
        ("Rep_Dauzere", "Dauzere"),
    ]
    for titulo_hoja, clave in repeticiones:
        ws = wb.create_sheet(titulo_hoja)
        filas = grupos.get(clave, {}).get("rep", [])
        _escribir_hoja(ws, COLS_REP, filas)

    wb.save(ruta_xlsx)


# EXPORTACION: CONFIGURACION Y HARDWARE

def _detectar_hardware():
    """Detecta informacion basica del entorno de ejecucion."""
    info = {}
    info["sistema"] = "%s %s" % (platform.system(), platform.release())
    info["maquina"] = platform.machine()
    info["procesador"] = platform.processor() or "no disponible"
    info["python"] = platform.python_version()
    info["nucleos_logicos"] = os.cpu_count() or "no disponible"

    # Modelo de CPU 
    modelo = None
    try:
        with open("/proc/cpuinfo", encoding="utf-8") as f:
            for linea in f:
                if linea.lower().startswith("model name"):
                    modelo = linea.split(":", 1)[1].strip()
                    break
    except OSError:
        pass
    info["cpu"] = modelo or info["procesador"]

    # Memoria total 
    ram = None
    try:
        with open("/proc/meminfo", encoding="utf-8") as f:
            for linea in f:
                if linea.startswith("MemTotal"):
                    kb = int(linea.split()[1])
                    ram = "%.1f GB" % (kb / (1024 * 1024))
                    break
    except OSError:
        pass
    info["ram"] = ram or "no disponible"
    return info


def exportar_configuracion(ruta_txt, config):
    """Genera configuracion.txt con parametros, semillas y hardware."""
    hw = _detectar_hardware()
    semillas = [config["semilla_base"] + r for r in range(config["n_reps"])]

    lineas = []
    sep = "=" * 60
    lineas.append(sep)
    lineas.append("CONFIGURACION DEL EXPERIMENTO")
    lineas.append(sep)
    lineas.append("")
    lineas.append("[Parametros]")
    lineas.append("  Fecha de ejecucion      : %s" % config["timestamp_legible"])
    lineas.append("  Metaheuristica          : Algoritmo Memetico (GA + Busqueda Local)")
    lineas.append("  Presupuesto por ejecucion: %d segundos" % config["tiempo_max"])
    lineas.append("  Repeticiones por instancia: %d" % config["n_reps"])
    lineas.append("  Semillas                : %s" %
                  ", ".join(str(s) for s in semillas))
    lineas.append("  Total de instancias     : %d" % config["total_instancias"])
    lineas.append("  Total de ejecuciones    : %d" % config["total_ejecuciones"])
    lineas.append("")
    lineas.append("[Benchmarks]")
    lineas.append("  Hurink (1994)      : edata, rdata, vdata x {la01,la06,la11,la16,la21}")
    lineas.append("  Brandimarte (1993) : Mk01..Mk10")
    lineas.append("  Dauzere (1998)     : 01a,04a,07a,10a,13a,16a")
    lineas.append("")
    lineas.append("[Hardware detectado]")
    lineas.append("  Sistema operativo  : %s" % hw["sistema"])
    lineas.append("  Arquitectura       : %s" % hw["maquina"])
    lineas.append("  CPU                : %s" % hw["cpu"])
    lineas.append("  Nucleos logicos    : %s" % hw["nucleos_logicos"])
    lineas.append("  Memoria RAM total  : %s" % hw["ram"])
    lineas.append("  Version de Python  : %s" % hw["python"])
    lineas.append("")

    with open(ruta_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas))


# PROGRAMA PRINCIPAL

def main():
    parser = argparse.ArgumentParser(
        description="Experimentos del Algoritmo Memetico para FJSP "
                    "(Hurink / Brandimarte / Dauzere).")
    parser.add_argument("--tiempo", type=float, default=60.0,
                        help="Segundos por ejecucion (def. 60).")
    parser.add_argument("--reps", type=int, default=5,
                        help="Repeticiones por instancia (def. 5).")
    parser.add_argument("--semilla-base", type=int, default=100,
                        help="Semilla inicial (def. 100).")
    parser.add_argument("--instancias-dir", default="instancias",
                        help="Carpeta raiz de instancias (def. instancias).")
    parser.add_argument("--salida-dir", default="resultados",
                        help="Carpeta de resultados (def. resultados).")
    args = parser.parse_args()

    os.makedirs(args.salida_dir, exist_ok=True)
    ahora = datetime.now()

    total_instancias = (len(HURINK_SUBCONJUNTOS) * len(HURINK_INSTANCIAS)
                        + len(BRANDIMARTE_INSTANCIAS)
                        + len(DAUZERE_INSTANCIAS))
    config = {
        "tiempo_max": args.tiempo,
        "n_reps": args.reps,
        "semilla_base": args.semilla_base,
        "timestamp_legible": ahora.strftime("%Y-%m-%d %H:%M:%S"),
        "total_instancias": total_instancias,
        "total_ejecuciones": total_instancias * args.reps,
    }

    print("=" * 70)
    print("EXPERIMENTOS MEMETICO-FJSP  |  %s" % config["timestamp_legible"])
    print("presupuesto=%ds/ejec  reps=%d  semillas=%d..%d"
          % (args.tiempo, args.reps, args.semilla_base,
             args.semilla_base + args.reps - 1))
    print("instancias=%d  ejecuciones=%d"
          % (total_instancias, config["total_ejecuciones"]))
    print("=" * 70)

    t0 = time.perf_counter()
    datos = correr_todo(args.instancias_dir, args.tiempo, args.reps,
                        args.semilla_base, registro=print)
    dur = (time.perf_counter() - t0) / 60

    ruta_txt = os.path.join(args.salida_dir, "resultados_reporte.txt")
    ruta_xlsx = os.path.join(args.salida_dir, "resultados.xlsx")
    ruta_cfg = os.path.join(args.salida_dir, "configuracion.txt")

    exportar_reporte_txt(ruta_txt, datos, config)
    exportar_xlsx(ruta_xlsx, datos)
    exportar_configuracion(ruta_cfg, config)

    print("\n" + "=" * 70)
    print("Experimento terminado en %.1f min." % dur)
    print("Archivos generados en %s/:" % args.salida_dir)
    print("  - resultados_reporte.txt")
    print("  - resultados.xlsx")
    print("  - configuracion.txt")
    print("=" * 70)


if __name__ == "__main__":
    main()
