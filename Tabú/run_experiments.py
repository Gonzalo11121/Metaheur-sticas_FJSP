# -*- coding: utf-8 -*-
"""
run_experiments.py
==================
Driver de experimentos para la Busqueda Tabu aplicada al FJSP.

Cubre tres familias de benchmarks:
  * Hurink, Jurisch y Thole (1994): subconjuntos edata, rdata y vdata
    (instancias la01, la06, la11, la16, la21).
  * Brandimarte (1993): Mk01 .. Mk10.
  * Dauzere-Peres y Paulli (1997): 01a, 04a, 07a, 10a, 13a, 16a.

Por cada instancia se ejecutan N_REPETICIONES con semillas fijas y un
presupuesto temporal uniforme. Se registran makespan, iteraciones y tiempo,
se calcula el gap% respecto al Upper Bound de referencia y se agregan las
metricas por instancia (mejor, promedio, desviacion estandar muestral y peor).

Salidas (carpeta resultados/):
  * resultados_reporte.txt  -> reporte legible organizado por benchmark.
  * resultados.xlsx         -> libro con resumen general, una hoja por
                               benchmark y una hoja de detalle por repeticion.
  * configuracion.txt       -> parametros, semillas y hardware detectado.

Uso:
    python3 run_experiments.py                  # corrida completa (60 s/ejecucion)
    python3 run_experiments.py --tiempo 5       # prueba rapida (5 s/ejecucion)
    python3 run_experiments.py --reps 3 --tiempo 10
"""

import argparse
import datetime
import math
import os
import platform
import sys

from fjsp_parser import parsear_archivo
from taboo_search import resolver, verificar_factibilidad

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ===========================================================================
#  Configuracion del experimento
# ===========================================================================
N_REPETICIONES = 5            # repeticiones por instancia
SEMILLA_BASE = 100            # semilla = SEMILLA_BASE + numero_de_repeticion (0..4)
LIMITE_TIEMPO = 60.0          # segundos por ejecucion (presupuesto uniforme)

INSTANCIAS_DIR = "instancias"
SALIDA_DIR = "resultados"


# ---------------------------------------------------------------------------
#  Upper Bounds (UB) de referencia para el calculo de gap%
# ---------------------------------------------------------------------------
UB_HURINK = {
    "edata": {"la01": 609, "la06": 800, "la11": 1071, "la16": 717, "la21": 835},
    "rdata": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 833},
    "vdata": {"la01": 570, "la06": 799, "la11": 1071, "la16": 717, "la21": 800},
}

UB_BRANDIMARTE = {
    "Mk01": 42, "Mk02": 32, "Mk03": 211, "Mk04": 81, "Mk05": 186,
    "Mk06": 86, "Mk07": 157, "Mk08": 523, "Mk09": 369, "Mk10": 296,
}

UB_DAUZERE = {
    "01a": 2530, "04a": 2565, "07a": 2408, "10a": 2362, "13a": 2302, "16a": 2301,
}


# ---------------------------------------------------------------------------
#  Definicion de los grupos de benchmark
#
#  Cada grupo declara:
#    clave        : identificador corto (tambien nombre de hoja en el xlsx)
#    benchmark    : etiqueta descriptiva para reportes
#    instancias   : lista ordenada de instancias
#    ruta         : funcion instancia -> ruta del archivo en disco
#    ub           : dict instancia -> Upper Bound
# ---------------------------------------------------------------------------
def _ruta_hurink(subset):
    def _f(inst):
        return os.path.join(INSTANCIAS_DIR, "Hurink_Data", subset, inst + ".txt")
    return _f


def _ruta_brandimarte(inst):
    return os.path.join(INSTANCIAS_DIR, "Brandimarte_Data", "Text", inst + ".fjs")


def _ruta_dauzere(inst):
    return os.path.join(INSTANCIAS_DIR, "dauzere", inst + ".fjs")


HURINK_INSTANCIAS = ["la01", "la06", "la11", "la16", "la21"]

GRUPOS = [
    {"clave": "Hurink_edata", "benchmark": "Hurink (edata)",
     "instancias": HURINK_INSTANCIAS, "ruta": _ruta_hurink("edata"),
     "ub": UB_HURINK["edata"]},
    {"clave": "Hurink_rdata", "benchmark": "Hurink (rdata)",
     "instancias": HURINK_INSTANCIAS, "ruta": _ruta_hurink("rdata"),
     "ub": UB_HURINK["rdata"]},
    {"clave": "Hurink_vdata", "benchmark": "Hurink (vdata)",
     "instancias": HURINK_INSTANCIAS, "ruta": _ruta_hurink("vdata"),
     "ub": UB_HURINK["vdata"]},
    {"clave": "Brandimarte", "benchmark": "Brandimarte",
     "instancias": ["Mk%02d" % k for k in range(1, 11)],
     "ruta": _ruta_brandimarte, "ub": UB_BRANDIMARTE},
    {"clave": "Dauzere", "benchmark": "Dauzere",
     "instancias": ["01a", "04a", "07a", "10a", "13a", "16a"],
     "ruta": _ruta_dauzere, "ub": UB_DAUZERE},
]


# ===========================================================================
#  Utilidades estadisticas
# ===========================================================================
def promedio(xs):
    return sum(xs) / len(xs)


def desviacion_estandar_muestral(xs):
    """Desviacion estandar muestral (n-1). Devuelve 0.0 si n<2."""
    n = len(xs)
    if n < 2:
        return 0.0
    m = promedio(xs)
    return math.sqrt(sum((x - m) ** 2 for x in xs) / (n - 1))


def _gap(valor, ub):
    return (100.0 * (valor - ub) / ub) if ub else None


def timestamp():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ===========================================================================
#  Experimento principal
# ===========================================================================
def correr_experimentos(limite_tiempo, n_reps):
    """Ejecuta todos los grupos y devuelve (detalle, resumen)."""
    detalle = []   # una fila por ejecucion
    resumen = []   # una fila por instancia

    total = sum(len(g["instancias"]) for g in GRUPOS)
    idx = 0
    for grupo in GRUPOS:
        clave = grupo["clave"]
        for inst_name in grupo["instancias"]:
            idx += 1
            ruta = grupo["ruta"](inst_name)
            etiqueta = "%s/%s" % (clave, inst_name)
            if not os.path.isfile(ruta):
                print("[%2d/%2d] %-22s  ARCHIVO NO ENCONTRADO (se omite)"
                      % (idx, total, etiqueta))
                continue

            inst = parsear_archivo(ruta, nombre=etiqueta)
            ub = grupo["ub"].get(inst_name)

            makespans, iters, tiempos = [], [], []
            print("[%2d/%2d] %-22s  (trabajos=%d, maquinas=%d, ops=%d)"
                  % (idx, total, etiqueta, inst.n_trabajos, inst.n_maquinas,
                     inst.n_operaciones))
            for rep in range(n_reps):
                semilla = SEMILLA_BASE + rep
                r = resolver(inst, semilla=semilla, limite_tiempo=limite_tiempo)
                factible, mk_chk, _msg = verificar_factibilidad(
                    inst, r["asignacion"], r["secuencia"])
                if (not factible) or (mk_chk != r["makespan"]):
                    print("      !! ADVERTENCIA: solucion no factible o inconsistente "
                          "(rep %d, semilla %d)" % (rep, semilla))
                makespans.append(r["makespan"])
                iters.append(r["iteraciones"])
                tiempos.append(r["tiempo"])
                gap_run = _gap(r["makespan"], ub)
                print("        rep %d (semilla %3d): Cmax=%6d  gap=%s  "
                      "iters=%7d  t=%5.1fs"
                      % (rep, semilla, r["makespan"],
                         ("%6.2f%%" % gap_run) if gap_run is not None else "   n/d",
                         r["iteraciones"], r["tiempo"]))
                detalle.append({
                    "grupo": clave, "benchmark": grupo["benchmark"],
                    "instancia": inst_name, "repeticion": rep, "semilla": semilla,
                    "makespan": r["makespan"], "gap_pct": gap_run,
                    "iteraciones": r["iteraciones"], "tiempo_s": r["tiempo"],
                })

            mejor = min(makespans)
            peor = max(makespans)
            prom = promedio(makespans)
            std = desviacion_estandar_muestral(makespans)
            fila = {
                "grupo": clave, "benchmark": grupo["benchmark"],
                "instancia": inst_name,
                "trabajos": inst.n_trabajos, "maquinas": inst.n_maquinas,
                "operaciones": inst.n_operaciones,
                "ub": ub, "mejor": mejor, "promedio": prom, "desv_std": std,
                "peor": peor, "gap_mejor": _gap(mejor, ub),
                "gap_prom": _gap(prom, ub),
                "iter_prom": promedio(iters), "tiempo_prom": promedio(tiempos),
            }
            resumen.append(fila)
            print("        => mejor=%d  prom=%.1f  std=%.2f  peor=%d  gap_mejor=%s"
                  % (mejor, prom, std, peor,
                     ("%.2f%%" % fila["gap_mejor"]) if fila["gap_mejor"] is not None
                     else "n/d"))
    return detalle, resumen


# ===========================================================================
#  Deteccion de hardware
# ===========================================================================
def _ram_total_gb():
    """Memoria fisica total en GB (mejor esfuerzo, multiplataforma)."""
    try:
        if hasattr(os, "sysconf") and "SC_PHYS_PAGES" in os.sysconf_names:
            paginas = os.sysconf("SC_PHYS_PAGES")
            tam = os.sysconf("SC_PAGE_SIZE")
            return paginas * tam / (1024 ** 3)
    except (ValueError, OSError):
        pass
    if platform.system() == "Windows":
        try:
            import ctypes

            class _MEM(ctypes.Structure):
                _fields_ = [("dwLength", ctypes.c_ulong),
                            ("dwMemoryLoad", ctypes.c_ulong),
                            ("ullTotalPhys", ctypes.c_ulonglong),
                            ("ullAvailPhys", ctypes.c_ulonglong),
                            ("ullTotalPageFile", ctypes.c_ulonglong),
                            ("ullAvailPageFile", ctypes.c_ulonglong),
                            ("ullTotalVirtual", ctypes.c_ulonglong),
                            ("ullAvailVirtual", ctypes.c_ulonglong),
                            ("ullAvailExtendedVirtual", ctypes.c_ulonglong)]
            estado = _MEM()
            estado.dwLength = ctypes.sizeof(_MEM)
            ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(estado))
            return estado.ullTotalPhys / (1024 ** 3)
        except Exception:
            return None
    return None


def _modelo_cpu():
    """Nombre comercial del procesador (mejor esfuerzo, multiplataforma)."""
    sistema = platform.system()
    if sistema == "Linux":
        try:
            with open("/proc/cpuinfo", "r") as f:
                for linea in f:
                    if linea.lower().startswith("model name"):
                        return linea.split(":", 1)[1].strip()
        except OSError:
            pass
    elif sistema == "Darwin":
        try:
            import subprocess
            salida = subprocess.check_output(
                ["sysctl", "-n", "machdep.cpu.brand_string"])
            return salida.decode().strip()
        except Exception:
            pass
    proc = platform.processor()
    return proc if proc else "(no informado)"


def detectar_hardware():
    ram = _ram_total_gb()
    return {
        "sistema": "%s %s" % (platform.system(), platform.release()),
        "arquitectura": platform.machine(),
        "procesador": _modelo_cpu(),
        "nucleos_logicos": os.cpu_count() or 0,
        "ram_gb": ram,
        "python": platform.python_version(),
        "implementacion": platform.python_implementation(),
    }


# ===========================================================================
#  Exportador: reporte de texto
# ===========================================================================
def exportar_reporte_txt(resumen, ruta, limite_tiempo, n_reps):
    enc = ("%-8s %5s %6s %9s %8s %6s %11s %11s %10s %9s"
           % ("inst", "UB", "mejor", "promedio", "std", "peor",
              "gap_mejor%", "gap_prom%", "iter_prom", "t_prom_s"))
    ancho = len(enc)
    lineas = []
    lineas.append("=" * ancho)
    lineas.append("  RESULTADOS BUSQUEDA TABU - FJSP")
    lineas.append("  Generado: %s" % timestamp())
    lineas.append("  Repeticiones por instancia: %d   |   Tiempo por ejecucion: %.0f s"
                  % (n_reps, limite_tiempo))
    lineas.append("  Semillas: %s"
                  % ", ".join(str(SEMILLA_BASE + r) for r in range(n_reps)))
    lineas.append("=" * ancho)

    def _fmt(v, dec=0):
        if v is None:
            return "n/d"
        if dec == 0:
            return str(int(round(v)))
        return ("%." + str(dec) + "f") % v

    for grupo in GRUPOS:
        filas = [f for f in resumen if f["grupo"] == grupo["clave"]]
        if not filas:
            continue
        lineas.append("")
        lineas.append("### %s" % grupo["benchmark"])
        lineas.append("-" * ancho)
        lineas.append(enc)
        lineas.append("-" * ancho)
        for f in filas:
            lineas.append("%-8s %5s %6s %9s %8s %6s %11s %11s %10s %9s" % (
                f["instancia"], _fmt(f["ub"]), _fmt(f["mejor"]),
                _fmt(f["promedio"], 1), _fmt(f["desv_std"], 2), _fmt(f["peor"]),
                _fmt(f["gap_mejor"], 2), _fmt(f["gap_prom"], 2),
                _fmt(f["iter_prom"], 1), _fmt(f["tiempo_prom"], 2)))
        gaps = [f["gap_mejor"] for f in filas if f["gap_mejor"] is not None]
        if gaps:
            lineas.append("-" * ancho)
            lineas.append("  gap_mejor%% promedio del grupo: %.2f%%" % promedio(gaps))

    lineas.append("")
    lineas.append("=" * ancho)
    lineas.append("Notas:")
    lineas.append(" - gap% = 100 * (makespan - UB) / UB.")
    lineas.append(" - std  = desviacion estandar muestral (n-1) de los makespans.")
    lineas.append(" - UB de referencia: Hurink (1994), Brandimarte y Dauzere segun")
    lineas.append("   Mastrolilli y Gambardella (2000).")

    texto = "\n".join(lineas) + "\n"
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(texto)
    return texto


# ===========================================================================
#  Exportador: libro de Excel
# ===========================================================================
_COL_RESUMEN_GEN = [
    ("Benchmark", "benchmark"), ("Instancia", "instancia"),
    ("Trabajos", "trabajos"), ("Maquinas", "maquinas"),
    ("Operaciones", "operaciones"), ("UB", "ub"), ("Mejor", "mejor"),
    ("Promedio", "promedio"), ("Desv_Std", "desv_std"), ("Peor", "peor"),
    ("Gap_Mejor%", "gap_mejor"), ("Gap_Prom%", "gap_prom"),
    ("Iter_Prom", "iter_prom"), ("Tiempo_Prom_s", "tiempo_prom"),
]
_COL_RESUMEN_GRUPO = _COL_RESUMEN_GEN[1:]   # sin la columna Benchmark
_COL_REP = [
    ("Instancia", "instancia"), ("Repeticion", "repeticion"),
    ("Semilla", "semilla"), ("Makespan", "makespan"), ("Gap%", "gap_pct"),
    ("Iteraciones", "iteraciones"), ("Tiempo_s", "tiempo_s"),
]

_DECIMALES = {
    "promedio": 2, "desv_std": 2, "gap_mejor": 2, "gap_prom": 2,
    "iter_prom": 1, "tiempo_prom": 2, "gap_pct": 2, "tiempo_s": 2,
}


def _valor_celda(fila, campo):
    v = fila.get(campo)
    if v is None:
        return "n/d"
    if campo in _DECIMALES and isinstance(v, float):
        return round(v, _DECIMALES[campo])
    return v


def _escribir_hoja(ws, columnas, filas):
    encabezado_fill = PatternFill("solid", fgColor="4C72B0")
    encabezado_font = Font(bold=True, color="FFFFFF")
    for c, (titulo, _campo) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=c, value=titulo)
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal="center")
    for r, fila in enumerate(filas, start=2):
        for c, (_titulo, campo) in enumerate(columnas, start=1):
            ws.cell(row=r, column=c, value=_valor_celda(fila, campo))
    for c, (titulo, _campo) in enumerate(columnas, start=1):
        ancho = max(len(str(titulo)) + 2, 11)
        ws.column_dimensions[get_column_letter(c)].width = ancho
    ws.freeze_panes = "A2"


def exportar_xlsx(resumen, detalle, ruta):
    if not _OPENPYXL:
        print("  (openpyxl no disponible: se omite resultados.xlsx)")
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen_General"
    _escribir_hoja(ws, _COL_RESUMEN_GEN, resumen)

    for grupo in GRUPOS:
        clave = grupo["clave"]
        filas = [f for f in resumen if f["grupo"] == clave]
        ws = wb.create_sheet(title=clave[:31])
        _escribir_hoja(ws, _COL_RESUMEN_GRUPO, filas)

    for grupo in GRUPOS:
        clave = grupo["clave"]
        filas = [f for f in detalle if f["grupo"] == clave]
        ws = wb.create_sheet(title=("Rep_" + clave)[:31])
        _escribir_hoja(ws, _COL_REP, filas)

    wb.save(ruta)
    return True


# ===========================================================================
#  Exportador: configuracion
# ===========================================================================
def exportar_configuracion(ruta, limite_tiempo, n_reps):
    hw = detectar_hardware()
    semillas = [SEMILLA_BASE + r for r in range(n_reps)]
    n_inst = sum(len(g["instancias"]) for g in GRUPOS)
    ram = "%.1f GB" % hw["ram_gb"] if hw["ram_gb"] else "(no detectada)"

    lineas = [
        "CONFIGURACION DEL EXPERIMENTO - BUSQUEDA TABU FJSP",
        "Generado: %s" % timestamp(),
        "",
        "[Parametros del experimento]",
        "  Metaheuristica           : Busqueda Tabu (Tabu Search)",
        "  Objetivo                 : minimizar makespan (Cmax)",
        "  Instancias totales       : %d" % n_inst,
        "  Repeticiones por instancia: %d" % n_reps,
        "  Tiempo por ejecucion (s) : %.0f" % limite_tiempo,
        "  Semilla base             : %d" % SEMILLA_BASE,
        "  Semillas utilizadas      : %s" % ", ".join(str(s) for s in semillas),
        "",
        "[Benchmarks evaluados]",
    ]
    for grupo in GRUPOS:
        lineas.append("  %-16s: %s"
                      % (grupo["benchmark"], ", ".join(grupo["instancias"])))

    lineas += [
        "",
        "[Parametros por defecto de la metaheuristica]",
        "  Tenencia tabu            : dinamica, escalada ~ sqrt(n_operaciones)",
        "  Vecindarios              : N1 reasignacion, N2 intercambio, N3 reubicacion",
        "  Maximo de vecinos/iter   : 300 (muestreo aleatorio si se supera)",
        "  Iter. sin mejora p/diver.: 40",
        "  Criterio de aspiracion   : se acepta movimiento tabu si mejora el mejor global",
        "  Solucion inicial         : multi-inicio con balanceo de carga",
        "",
        "[Hardware detectado]",
        "  Sistema operativo        : %s" % hw["sistema"],
        "  Arquitectura             : %s" % hw["arquitectura"],
        "  Procesador               : %s" % hw["procesador"],
        "  Nucleos logicos          : %s" % hw["nucleos_logicos"],
        "  Memoria RAM total        : %s" % ram,
        "",
        "[Entorno de ejecucion]",
        "  Python                   : %s (%s)" % (hw["python"], hw["implementacion"]),
        "  openpyxl disponible      : %s" % ("si" if _OPENPYXL else "no"),
    ]
    texto = "\n".join(lineas) + "\n"
    with open(ruta, "w", encoding="utf-8") as f:
        f.write(texto)
    return texto


# ===========================================================================
#  Programa principal
# ===========================================================================
def main():
    ap = argparse.ArgumentParser(description="Experimentos Tabu Search FJSP")
    ap.add_argument("--tiempo", type=float, default=LIMITE_TIEMPO,
                    help="segundos por ejecucion (def. %.0f)" % LIMITE_TIEMPO)
    ap.add_argument("--reps", type=int, default=N_REPETICIONES,
                    help="repeticiones por instancia (def. %d)" % N_REPETICIONES)
    args = ap.parse_args()

    n_inst = sum(len(g["instancias"]) for g in GRUPOS)
    est_total = n_inst * args.reps * args.tiempo
    print("Experimento Tabu Search - FJSP")
    print("Instancias: %d  |  repeticiones: %d  |  tiempo/ejecucion: %.0f s"
          % (n_inst, args.reps, args.tiempo))
    print("Tiempo total estimado: ~%.0f s (~%.1f min)\n"
          % (est_total, est_total / 60.0))

    detalle, resumen = correr_experimentos(args.tiempo, args.reps)

    if not resumen:
        print("\nNo se proceso ninguna instancia. Revisa la carpeta '%s'."
              % INSTANCIAS_DIR)
        return 1

    os.makedirs(SALIDA_DIR, exist_ok=True)
    ruta_txt = os.path.join(SALIDA_DIR, "resultados_reporte.txt")
    ruta_xlsx = os.path.join(SALIDA_DIR, "resultados.xlsx")
    ruta_cfg = os.path.join(SALIDA_DIR, "configuracion.txt")

    texto = exportar_reporte_txt(resumen, ruta_txt, args.tiempo, args.reps)
    ok_xlsx = exportar_xlsx(resumen, detalle, ruta_xlsx)
    exportar_configuracion(ruta_cfg, args.tiempo, args.reps)

    print("\n" + texto)
    print("Archivos generados:")
    print("  - %s" % ruta_txt)
    if ok_xlsx:
        print("  - %s" % ruta_xlsx)
    print("  - %s" % ruta_cfg)
    return 0


if __name__ == "__main__":
    sys.exit(main())
